import tkinter as tk
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from core.config_manager import ConfigManager

class WeightOrdersExporter:
    """Экспортер данных в Excel файл для весовых заказов"""

    def __init__(self, excel_file_path, roll_module, preview_module, coordinator=None):
        self.excel_file_path = excel_file_path
        self.roll_module = roll_module
        self.preview_module = preview_module
        self.config_manager = ConfigManager()
        self.coordinator = coordinator
        # Подписываемся на координатор если он есть
        if self.coordinator and hasattr(self.coordinator, 'subscribe'):
            self.coordinator.subscribe(self.on_settings_changed)
            
        self.wb = None
        self.ws = None
        
    def export_data(self, enable_pallet=False, pallet_data=None, multitype_mode=False):
        """Основной метод экспорта данных с четким разделением по цехам"""
        try:
            # Используем правильный путь к файлу в зависимости от цеха
            actual_file_path = self.get_excel_file_path()
            
            if not os.path.exists(actual_file_path):
                raise FileNotFoundError(f"Файл не найден: {actual_file_path}")
            
            # Загружаем книгу
            self.wb = load_workbook(actual_file_path)
            
            # Проверяем изготовителя
            self._update_manufacturer_info()
            
            # Определяем цех
            is_second_file = self._is_second_file()
            
            # Ветвление по цехам и режимам
            if multitype_mode:                             
                # Режим много-видового листа
                if is_second_file:
                    # Цех 2 + много видов
                    self.ws = self.wb["Много видов"]
                    result = self._export_to_multitype_sheet_for_second_workshop(pallet_data)
                    return result
                else:
                    # Цех 1 + много видов
                    self.ws = self.wb["Лист много видов"]
                    result = self._export_to_multitype_sheet_first_workshop(pallet_data)
                    return result
            
            # Стандартный режим (коробка/поддон)
            if is_second_file:
                # Цех 2
                if enable_pallet:
                    # Цех 2 + поддон:
                    result = self.export_to_pallet_list_sheet()
                    self.wb.close()
                    return result
                else:
                    # Цех 2 + коробка
                    self.ws = self.wb["Поддон"]
                    self._export_basic_info_for_second_workshop()
                    self._export_box_data_for_second_workshop()
                    all_fitted = self._export_rolls_to_empty_cells_for_second_workshop()
            else:
                # Цех 1
                if enable_pallet:
                    # Цех 1 + поддон
                    self.ws = self.wb["Лист для паллеты"]
                    self._export_basic_info_for_first_workshop()
                    all_fitted = self._export_pallet_data_for_first_workshop(pallet_data)
                else:
                    # Цех 1 + коробка
                    self.ws = self.wb["Лист для коробки"]
                    self._export_basic_info_for_first_workshop()
                    self._export_box_data_for_first_workshop()
                    all_fitted = self._export_rolls_to_empty_cells_for_first_workshop()
            
            # Сохраняем изменения
            self.wb.save(actual_file_path)
            
            # УНИВЕРСАЛЬНОЕ уведомление для ЛЮБОГО успешного экспорта
            if self.coordinator and hasattr(self.coordinator, 'notify'):
                sheet_name = self.ws.title if self.ws else "Неизвестный лист"
                self.coordinator.notify("excel_exported", {
                    'file_path': actual_file_path,
                    'sheet_name': sheet_name,
                    'enable_pallet': enable_pallet,
                    'multitype_mode': multitype_mode,
                    'workshop': '2' if is_second_file else '1'
                })
            
            # Возвращаем словарь с результатами
            return {
                'success': True,
                'all_fitted': all_fitted
            }
            
        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            return {'success': False, 'error': str(e)}
        finally:
            if self.wb:
                self.wb.close()

    # ==================== МЕТОДЫ ДЛЯ ЦЕХА 1 ====================

    def _export_basic_info_for_first_workshop(self, skip_product_name=False):
        """Экспортирует основную информацию для цеха 1"""
        try:
            # Заказчик - D5
            if self.roll_module and hasattr(self.roll_module, 'customer_var'):
                customer = self.roll_module.customer_var.get()
                self._set_cell_value('D5', customer)
            
            # Тип упаковки - D6
            if self.roll_module and hasattr(self.roll_module, 'box_size_var'):
                box_type = self.roll_module.box_size_var.get()
                self._set_cell_value('D6', box_type)
            
            # Номер заказа - D8
            if self.roll_module:
                order_prefix = getattr(self.roll_module, 'order_prefix', None).get()
                order_number = getattr(self.roll_module, 'order_number', None).get()
                order_suffix = getattr(self.roll_module, 'order_suffix', None).get()
                full_order = f"{order_prefix}{order_number}{order_suffix}"
                self._set_cell_value('D8', full_order)
            
            # Наименование продукции - D10
            if not skip_product_name and self.roll_module and hasattr(self.roll_module, 'product_text'):
                product_text = self.roll_module.product_text.get("1.0", "end-1c").strip()
                self._set_cell_value('D10', product_text)
                cell = self.ws['D10']
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Дата упаковки - F37
            if self.roll_module and hasattr(self.roll_module, 'date_var'):
                date = self.roll_module.date_var.get()
                self._set_cell_value('F37', date)
            
            # Упаковщик - E41
            if self.roll_module and hasattr(self.roll_module, 'packer_var'):
                packer = self.roll_module.packer_var.get()
                self._set_cell_value('E41', packer)
                
            # Тип продукта - E39 (добавляем)
            if self.roll_module and hasattr(self.roll_module, 'product_type_var'):
                product_type = self.roll_module.product_type_var.get()
                self._set_cell_value('E39', product_type)
            
            # TU номер - A39 (добавляем)
            tu_number = self._get_tu_number()  # используем тот же метод
            self._set_cell_value('A39', tu_number)
                    
        except Exception as e:
            print(f"Ошибка при экспорте базовой информации для цеха 1: {e}")

    def _export_box_data_for_first_workshop(self):
        """Экспортирует данные коробки для цеха 1"""
        if not self.roll_module:
            return True
            
        try:
            # Вес коробки - K2
            box_weight = self._convert_to_number_if_possible(self.roll_module.box_weight_var.get())
            if box_weight:
                self._set_cell_value('K2', box_weight)
            
            return True
            
        except Exception as e:
            print(f"Ошибка при экспорте данных коробки для цеха 1: {e}")
            return True

    def _export_pallet_data_for_first_workshop(self, pallet_data):
        """Экспортирует данные поддона для цеха 1"""
        try:
            # Тип упаковки (поддон) - D6
            if pallet_data and pallet_data.get("pallet_type"):
                self._set_cell_value('D6', pallet_data["pallet_type"])
            
            # Вес поддона - K2
            if pallet_data and pallet_data.get("pallet_weight"):
                pallet_weight = self._convert_to_number_if_possible(pallet_data["pallet_weight"])
                self._set_cell_value('K2', pallet_weight)
            
            # Экспортируем данные коробок и возвращаем информацию о переполнении
            return self._export_boxes_to_empty_cells_for_first_workshop(pallet_data)
                
        except Exception as e:
            print(f"Ошибка при экспорте данных поддона для цеха 1: {e}")
            return True

    def _export_boxes_to_empty_cells_for_first_workshop(self, pallet_data):
        """Экспортирует данные коробок в пустые ячейки для поддона в цехе 1"""
        if not self.roll_module:
            return True
            
        try:
            # Получаем количество коробок из pallet_data
            boxes_count = 0
            if pallet_data and pallet_data.get("boxes_count"):
                try:
                    boxes_count_str = pallet_data["boxes_count"]
                    boxes_count = int(boxes_count_str) if boxes_count_str else 0
                except (ValueError, TypeError):
                    boxes_count = 0
            
            if boxes_count == 0:
                boxes_count = 1
            
            # Получаем данные одной коробки из roll_module
            gross_weight = self._convert_to_number_if_possible(self.roll_module.total_gross_var.get())
            net_weight = self._convert_to_number_if_possible(self.roll_module.total_net_var.get())
            quantity = self._convert_to_number_if_possible(self.roll_module.total_quantity_var.get())
            
            # Логика для 1 цеха
            empty_cells_found = 0
            
            # Заполняем левую часть (B14-B28, C14-C28, D14-D28)
            for row in range(14, 29):
                if empty_cells_found >= boxes_count:
                    break
                    
                if (self.ws[f'B{row}'].value is None and 
                    self.ws[f'C{row}'].value is None and 
                    self.ws[f'D{row}'].value is None):
                    
                    if gross_weight:
                        self._set_cell_value(f'B{row}', gross_weight)
                    if net_weight:
                        self._set_cell_value(f'C{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'D{row}', quantity)
                    empty_cells_found += 1
            
            # Заполняем правую часть (F14-F28, G14-G28, H14-H28)
            for row in range(14, 29):
                if empty_cells_found >= boxes_count:
                    break
                    
                if (self.ws[f'F{row}'].value is None and 
                    self.ws[f'G{row}'].value is None and 
                    self.ws[f'H{row}'].value is None):
                    
                    if gross_weight:
                        self._set_cell_value(f'F{row}', gross_weight)
                    if net_weight:
                        self._set_cell_value(f'G{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'H{row}', quantity)
                    empty_cells_found += 1
            
            return empty_cells_found >= boxes_count
                
        except Exception as e:
            print(f"Ошибка при экспорте данных коробок для цеха 1: {e}")
            return True

    def _export_rolls_to_empty_cells_for_first_workshop(self):
        """Экспортирует данные роликов в пустые ячейки для цеха 1"""
        if not self.roll_module:
            return True
            
        try:
            # Получаем количество роликов
            rolls_count = 0
            if self.roll_module and hasattr(self.roll_module, 'rolls_count_var'):
                try:
                    rolls_count_str = self.roll_module.rolls_count_var.get()
                    rolls_count = int(rolls_count_str) if rolls_count_str else 0
                except (ValueError, TypeError):
                    rolls_count = 0
            
            if rolls_count == 0:
                rolls_count = 1
            
            # Получаем данные одного ролика
            gross_weight = self._convert_to_number_if_possible(self.roll_module.gross_weight_kg_var.get())
            net_weight = self._convert_to_number_if_possible(self.roll_module.net_weight_kg_var.get())
            quantity = self._convert_to_number_if_possible(self.roll_module.quantity_var.get())
            
            # Находим пустые ячейки и заполняем их
            empty_cells_found = 0
            
            # Для первого файла - старый код
            for row in range(14, 29):
                if empty_cells_found >= rolls_count:
                    break
                    
                if (self.ws[f'B{row}'].value is None and
                    self.ws[f'C{row}'].value is None and
                    self.ws[f'D{row}'].value is None):
                    
                    if gross_weight:
                        self._set_cell_value(f'B{row}', gross_weight)
                    if net_weight:
                        self._set_cell_value(f'C{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'D{row}', quantity)
                    
                    empty_cells_found += 1
            
            for row in range(14, 29):
                if empty_cells_found >= rolls_count:
                    break
                    
                if (self.ws[f'F{row}'].value is None and 
                    self.ws[f'G{row}'].value is None and 
                    self.ws[f'H{row}'].value is None):
                    
                    if gross_weight:
                        self._set_cell_value(f'F{row}', gross_weight)
                    if net_weight:
                        self._set_cell_value(f'G{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'H{row}', quantity)
                    
                    empty_cells_found += 1
            
            return empty_cells_found >= rolls_count
                
        except Exception as e:
            print(f"Ошибка при экспорте данных роликов для цеха 1: {e}")
            return True

    # ==================== МЕТОДЫ ДЛЯ ЦЕХА 2 ====================

    def _export_basic_info_for_second_workshop(self, skip_product_name=False):
        """Экспортирует основную информацию для цеха 2"""
        try:
            # Заказчик - D7
            if self.roll_module and hasattr(self.roll_module, 'customer_var'):
                customer = self.roll_module.customer_var.get()
                self._set_cell_value('D7', customer)
                cell = self.ws['D7']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Тип упаковки - D3
            if self.roll_module and hasattr(self.roll_module, 'box_size_var'):
                box_type = self.roll_module.box_size_var.get()
                self._set_cell_value('D3', box_type)
                cell = self.ws['D3']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Номер заказа - D6
            if self.roll_module:
                order_prefix = getattr(self.roll_module, 'order_prefix', None).get()
                order_number = getattr(self.roll_module, 'order_number', None).get()
                order_suffix = getattr(self.roll_module, 'order_suffix', None).get()
                full_order = f"{order_prefix}{order_number}{order_suffix}"
                self._set_cell_value('D6', full_order)
                cell = self.ws['D6']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Наименование продукции - D8
            if not skip_product_name and self.roll_module and hasattr(self.roll_module, 'product_text'):
                product_text = self.roll_module.product_text.get("1.0", "end-1c").strip()
                self._set_cell_value('D8', product_text)
                cell = self.ws['D8']
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Дата упаковки - D37
            if self.roll_module and hasattr(self.roll_module, 'date_var'):
                date = self.roll_module.date_var.get()
                self._set_cell_value('D37', date)
            
            # Упаковщик - E41
            if self.roll_module and hasattr(self.roll_module, 'packer_var'):
                packer = self.roll_module.packer_var.get()
                self._set_cell_value('E41', packer)
            
            # Тип продукта - E39
            if self.roll_module and hasattr(self.roll_module, 'product_type_var'):
                product_type = self.roll_module.product_type_var.get()
                self._set_cell_value('E39', product_type)
            
            # TU номер - A39
            tu_number = self._get_tu_number()
            self._set_cell_value('A39', tu_number)
                    
        except Exception as e:
            print(f"Ошибка при экспорте базовой информации для цеха 2: {e}")

    def _export_box_data_for_second_workshop(self):
        """Экспортирует данные коробки для цеха 2"""
        if not self.roll_module:
            return True
            
        try:
            # Вес коробки - H3
            box_weight = self._convert_to_number_if_possible(self.roll_module.box_weight_var.get())
            if box_weight:
                self._set_cell_value('H3', box_weight)
            
            # Вес втулки (конвертация из граммов в кг) - D4
            sleeve_weight_g = self._convert_to_number_if_possible(self.roll_module.sleeve_weight_var.get())
            if sleeve_weight_g:
                sleeve_weight_kg = sleeve_weight_g / 1000
                self._set_cell_value('D4', sleeve_weight_kg)
            
            # Диаметр втулки - G4
            sleeve_diameter = self._convert_to_number_if_possible(self.roll_module.sleeve_diameter_var.get())
            if sleeve_diameter:
                self._set_cell_value('G4', sleeve_diameter)
            
            return True
            
        except Exception as e:
            print(f"Ошибка при экспорте данных коробки для цеха 2: {e}")
            return True

    def _export_rolls_to_empty_cells_for_second_workshop(self):
        """Экспортирует данные роликов в пустые ячейки для цеха 2"""
        if not self.roll_module:
            return True
            
        try:
            # Получаем количество роликов
            rolls_count = 0
            if self.roll_module and hasattr(self.roll_module, 'rolls_count_var'):
                try:
                    rolls_count_str = self.roll_module.rolls_count_var.get()
                    rolls_count = int(rolls_count_str) if rolls_count_str else 0
                except (ValueError, TypeError):
                    rolls_count = 0
            
            if rolls_count == 0:
                rolls_count = 1
            
            # Получаем данные одного ролика
            net_weight = self._convert_to_number_if_possible(self.roll_module.net_weight_kg_var.get())
            quantity = self._convert_to_number_if_possible(self.roll_module.quantity_var.get())
            roll_length = self._convert_to_number_if_possible(self.roll_module.roll_length.get())
            
            # Находим пустые ячейки и заполняем их
            empty_cells_found = 0
            
            # Заполняем первую колонку (B и C) - ролики 1-20
            for row in range(10, 30):
                if empty_cells_found >= rolls_count:
                    break
                    
                # Проверяем только пару B,C
                if (self.ws[f'B{row}'].value is None and 
                    self.ws[f'C{row}'].value is None):
                    
                    # Соответствующий слот в L: для строки 10 -> L10, строки 11 -> L11 и т.д.
                    l_slot = row  # для первой колонки
                    
                    if net_weight:
                        self._set_cell_value(f'B{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'C{row}', quantity)
                    if roll_length:
                        self._set_cell_value(f'L{l_slot}', roll_length)
                    
                    empty_cells_found += 1

            # Заполняем вторую колонку (E и F) - ролики 21-40  
            for row in range(10, 30):
                if empty_cells_found >= rolls_count:
                    break
                    
                # Проверяем только пару E,F
                if (self.ws[f'E{row}'].value is None and 
                    self.ws[f'F{row}'].value is None):
                    
                    # Соответствующий слот в L: для строки 10 -> L30, строки 11 -> L31 и т.д.
                    l_slot = row + 20  # для второй колонки
                    
                    if net_weight:
                        self._set_cell_value(f'E{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'F{row}', quantity)
                    if roll_length:
                        self._set_cell_value(f'L{l_slot}', roll_length)
                    
                    empty_cells_found += 1

            # Заполняем третью колонку (H и I) - ролики 41-60
            for row in range(10, 30):
                if empty_cells_found >= rolls_count:
                    break
                    
                # Проверяем только пару H,I
                if (self.ws[f'H{row}'].value is None and 
                    self.ws[f'I{row}'].value is None):
                    
                    # Соответствующий слот в L: для строки 10 -> L50, строки 11 -> L51 и т.д.
                    l_slot = row + 40  # для третьей колонки
                    
                    if net_weight:
                        self._set_cell_value(f'H{row}', net_weight)
                    if quantity:
                        self._set_cell_value(f'I{row}', quantity)
                    if roll_length:
                        self._set_cell_value(f'L{l_slot}', roll_length)
                    
                    empty_cells_found += 1
            
            return empty_cells_found >= rolls_count
                
        except Exception as e:
            print(f"Ошибка при экспорте данных роликов для цеха 2: {e}")
            return True

    # ==================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ====================
    
    def export_to_pallet_list_sheet(self, excel_path=None, sheet_name=None):
        """
        Экспорт данных поддона в лист 'Список поддонов' без архивации
        """
        try:
            if excel_path is None:
                excel_path = self.get_excel_file_path()
            
            if sheet_name is None:
                sheet_name = "Список поддонов"
            
            workbook = load_workbook(excel_path)
            
            # Проверяем листы
            if "Поддон" not in workbook.sheetnames:
                workbook.close()
                return {'success': False, 'error': 'Лист "Поддон" не найден', 'all_fitted': False}
            
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {'success': False, 'error': f'Лист "{sheet_name}" не найден', 'all_fitted': False}
            
            pallet_sheet = workbook["Поддон"]
            list_sheet = workbook[sheet_name]
            
            # 1. Найти свободную строку (10-29)
            target_row = None
            for row in range(10, 30):
                if (list_sheet[f'D{row}'].value is None and 
                    list_sheet[f'F{row}'].value is None and
                    list_sheet[f'H{row}'].value is None and
                    list_sheet[f'L{row}'].value is None):
                    target_row = row
                    break
            
            if target_row is None:
                workbook.close()
                return {'success': False, 'error': f'Лист "{sheet_name}" переполнен (нет свободных строк 10-29)', 'all_fitted': False}
            
            # 2. Сохраняем текущие листы
            original_ws = getattr(self, 'ws', None)
            original_wb = getattr(self, 'wb', None)
            original_sheet_name = getattr(self, 'sheet_name', None)
            
            # 3. Временно переключаемся на лист списка
            self.wb = workbook
            self.ws = list_sheet
            self.sheet_name = sheet_name
            
            # 4. Копируем базовую информацию из "Поддона" в "Список поддонов"
            # Заказчик
            list_sheet["D7"] = pallet_sheet["D7"].value
            # Тип упаковки
            list_sheet["D3"] = pallet_sheet["D3"].value
            # Номер заказа
            list_sheet["D6"] = pallet_sheet["D6"].value
            # Изделие
            list_sheet["D8"] = pallet_sheet["D8"].value
            # Дата упаковки
            list_sheet["D37"] = pallet_sheet["D37"].value
            # Упаковщик
            list_sheet["E41"] = pallet_sheet["E41"].value
            # Вес втулки (кг)
            list_sheet["D4"] = pallet_sheet["D4"].value
            # Диаметр втулки
            list_sheet["G4"] = pallet_sheet["G4"].value
            # Тип продукта
            list_sheet["E39"] = pallet_sheet["E39"].value
            # TU номер
            list_sheet["A39"] = pallet_sheet["A39"].value
            
            # 5. Вычисляем и заполняем данные поддона
            totals = self._calculate_pallet_totals(pallet_sheet)
            
            # D - Кол-во роликов
            list_sheet[f'D{target_row}'] = totals['rolls_count']
            
            # F - Вес нетто
            list_sheet[f'F{target_row}'] = totals['total_weight']
            
            # H - Кол-во этикеток
            list_sheet[f'H{target_row}'] = totals['total_quantity']
            
            # L - Сумма длин
            list_sheet[f'L{target_row}'] = totals['total_length']
            
            # 6. Вес поддона (рассчитывается автоматически формулой)
            # В ячейке H3 уже должна быть формула
            
            # 7. Сохраняем файл
            workbook.save(excel_path)
            workbook.close()
            
            # 8. Восстанавливаем исходные листы
            if original_wb:
                self.wb = original_wb
            if original_ws:
                self.ws = original_ws
            if original_sheet_name:
                self.sheet_name = original_sheet_name
            
            return {
                'success': True, 
                'row_used': target_row,
                'total_weight': totals['total_weight'],
                'total_quantity': totals['total_quantity'],
                'all_fitted': True
            }
            
        except Exception as e:
            print(f"Ошибка экспорта в список поддонов: {e}")
            return {'success': False, 'error': str(e), 'all_fitted': False}

    def _calculate_pallet_totals(self, pallet_sheet):
        """Вычисляет итоги по поддону из листа 'Поддон'"""
        total_quantity = 0
        total_weight = 0
        total_length = 0
        rolls_count = 0
        
        # Пары колонок и соответствующие смещения для длины в L
        column_pairs = [
            ('B', 'C', 0),   # B,C - длина в L с тем же номером строки
            ('E', 'F', 20),  # E,F - длина в L со смещением +20
            ('H', 'I', 40)   # H,I - длина в L со смещением +40
        ]
        
        for weight_col, qty_col, l_offset in column_pairs:
            for row in range(10, 30):  # строки 10-29
                weight = pallet_sheet[f'{weight_col}{row}'].value
                quantity = pallet_sheet[f'{qty_col}{row}'].value
                
                if weight is not None or quantity is not None:
                    rolls_count += 1
                    
                    if weight is not None:
                        total_weight += weight
                    if quantity is not None:
                        total_quantity += quantity
                    
                    # Длина из столбца L
                    length_row = row + l_offset
                    length = pallet_sheet[f'L{length_row}'].value
                    if length is not None:
                        total_length += length
        
        return {
            'rolls_count': rolls_count,
            'total_weight': total_weight,
            'total_quantity': total_quantity,
            'total_length': total_length
        }
    
    def get_excel_file_path(self):
        """Возвращает путь к Excel файлу в зависимости от цеха и настроек"""
        try:
            # Получаем путь из настроек (shared_utils.json)
            settings = self.config_manager.load_json_settings("shared_utils.json")
            excel_folder = settings.get("weight_orders_xlsx", "")
            
            if not excel_folder:
                # Если папка не выбрана в настройках, используем исходный путь
                excel_folder = os.path.dirname(self.excel_file_path)
            
            # Определяем имя файла в зависимости от цеха
            workshop = "1"
            if hasattr(self, 'coordinator') and self.coordinator:
                workshop = self.coordinator.get_workshop()
            
            filename = "weight_orders.xlsx" if workshop == "1" else "weight_orders_2.xlsx"
            full_path = os.path.join(excel_folder, filename)
            
            # Проверяем существование файла
            if not os.path.exists(full_path):
                # Копируем из assets
                assets_file = self.config_manager.get_asset_path(filename)
                if os.path.exists(assets_file):
                    import shutil
                    shutil.copy2(assets_file, full_path)
                    print(f"Файл {filename} скопирован в {full_path}")
                else:
                    raise FileNotFoundError(f"Файл {filename} не найден в assets")
            
            return full_path
            
        except Exception as e:
            print(f"Ошибка получения пути к Excel файлу: {e}")
            # Fallback: возвращаем исходный путь
            return self.excel_file_path

    def _copy_excel_file_from_assets(self, assets_filename, target_path):
        """Копирует файл Excel из assets в целевую папку"""
        try:
            assets_file = self.config_manager.get_asset_path(assets_filename)
            
            if os.path.exists(assets_file):
                import shutil
                shutil.copy2(assets_file, target_path)
                print(f"Файл {assets_filename} скопирован из assets в {target_path}")
            else:
                raise FileNotFoundError(f"Файл {assets_filename} не найден в assets")
                
        except Exception as e:
            print(f"Ошибка копирования файла {assets_filename}: {e}")
            raise
        
    def _is_second_file(self):
        """Проверяет, является ли файл вторым файлом (weight_orders_2.xlsx)"""
        file_path = self.get_excel_file_path()
        if not file_path:
            return False
        # Просто проверяем имя файла, координатор уже учтен в get_excel_file_path
        return "weight_orders_2.xlsx" in os.path.basename(file_path)
    
    def clear_all_rolls(self, enable_pallet=False, multitype_mode=False):
        """Основной метод очистки - вызывает соответствующие под-методы"""
        try:
            actual_file_path = self.get_excel_file_path()
            
            if not os.path.exists(actual_file_path):
                raise FileNotFoundError(f"Файл не найден: {actual_file_path}")
            
            # Проверяем, не открыт ли файл в Excel
            if self._is_file_locked(actual_file_path):
                raise PermissionError(f"Файл {actual_file_path} открыт в Excel. Закройте его и попробуйте снова.")
            
            self.wb = load_workbook(actual_file_path)
            
            is_second_file = self._is_second_file()
            
            # Ветвление по режимам
            if multitype_mode:
                # Много-видовой режим
                if is_second_file:
                    # Цех 2: очищаем "Много видов"
                    self._clear_multitype_sheet_second_workshop()
                else:
                    # Цех 1: очищаем "Лист много видов"
                    self._clear_multitype_sheet_first_workshop()
            else:
                # Стандартный режим (коробка/поддон)
                if is_second_file:
                    if enable_pallet:
                        # Для 2 цеха, режим поддона: очищаем "Список поддонов"
                        self._clear_pallet_list_sheet()
                    else:
                        # Для 2 цеха, режим коробки: очищаем только "Поддон"
                        self._clear_pallet_sheet_second_workshop()
                else:
                    # Для 1 цеха
                    if enable_pallet:
                        self._clear_pallet_sheet_first_workshop()
                    else:
                        self._clear_box_sheet_first_workshop()
            
            # Сохраняем изменения
            self.wb.save(actual_file_path)
            
            # УНИВЕРСАЛЬНОЕ уведомление для ЛЮБОЙ успешной очистки
            if self.coordinator and hasattr(self.coordinator, 'notify'):
                self.coordinator.notify("excel_cleared", {
                    'file_path': actual_file_path,
                    'enable_pallet': enable_pallet,
                    'multitype_mode': multitype_mode,
                    'workshop': '2' if is_second_file else '1'
                })
            return True
            
        except PermissionError as e:
            print(f"Ошибка доступа к файлу: {e}")
            return False
        except FileNotFoundError as e:
            print(f"Файл не найден: {e}")
            return False
        except Exception as e:
            print(f"Ошибка очистки Excel: {e}")
            return False
        finally:
            if self.wb:
                self.wb.close()
                
    def _clear_multitype_sheet_first_workshop(self):
        """Очищает лист 'Лист много видов' для цеха 1"""
        if "Лист много видов" not in self.wb.sheetnames:
            return False
            
        self.ws = self.wb["Лист много видов"]
        
        # 1. Очищаем базовую информацию (как в clear_multitype_sheet)
        basic_info_cells = ['D5', 'D6', 'D8', 'F37', 'E41', 'K2', 'E39', 'A39']
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")
        
        # 2. Очищаем данные продуктов (строки 11-28)
        for row in range(11, 29):
            for col in ['A', 'B', 'F', 'G', 'H']:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        return True

    def _clear_multitype_sheet_second_workshop(self):
        """Очищает лист 'Много видов' для цеха 2"""
        if "Много видов" not in self.wb.sheetnames:
            return False
            
        self.ws = self.wb["Много видов"]
        
        # 1. Очищаем базовую информацию
        basic_info_cells = [
            'D7', 'D3', 'D6', 'D37', 'E41', 'H3', 
            'D4', 'G4', 'E39', 'A39'
        ]
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")
        
        # 2. Очищаем данные продуктов (строки 10-29)
        for row in range(10, 30):
            for col in ['A', 'B', 'H', 'I', 'L']:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        return True

    def _clear_pallet_sheet_second_workshop(self):
        """Очищает лист 'Поддон' для 2 цеха"""
        sheet_name = "Поддон"
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
            
        self.ws = self.wb[sheet_name]
        
        # Очищаем колонки коробок (B,C,E,F,H,I)
        rows_range = range(10, 30)  # строки 10-29
        weight_quantity_cols = ['B', 'C', 'E', 'F', 'H', 'I']
        
        for row in rows_range:
            for col in weight_quantity_cols:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        # Очищаем столбец L для 60 роликов (строки 10-69)
        for row in range(10, 70):
            try:
                cell = self.ws[f'L{row}']
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку L{row}: {e}")
        
        # Очищаем базовую информацию
        basic_info_cells = ['A1', 'D7', 'D3', 'D6', 'D8', 'D37', 'E41', 'H3', 
                           'D4', 'G4', 'E39', 'A39', 'D5']
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")

    def _clear_pallet_sheet_first_workshop(self):
        """Очищает лист 'Лист для паллеты' для 1 цеха"""
        sheet_name = "Лист для паллеты"
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
            
        self.ws = self.wb[sheet_name]
        
        # Очищаем данные
        rows_range = range(14, 29)  # строки 14-28
        cols_to_clear = ['B', 'C', 'D', 'F', 'G', 'H']
        
        for row in rows_range:
            for col in cols_to_clear:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        # Очищаем базовую информацию
        basic_info_cells = ['D5', 'D6', 'D8', 'D10', 'F37', 'E41', 'K2', 'E39', 'A39']
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")

    def _clear_box_sheet_first_workshop(self):
        """Очищает лист 'Лист для коробки' для 1 цеха"""
        sheet_name = "Лист для коробки"
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
            
        self.ws = self.wb[sheet_name]
        
        # Очищаем данные
        rows_range = range(14, 29)  # строки 14-28
        cols_to_clear = ['B', 'C', 'D', 'F', 'G', 'H']
        
        for row in rows_range:
            for col in cols_to_clear:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        # Очищаем базовую информацию
        basic_info_cells = ['B1', 'D5', 'D6', 'D8', 'D10', 'F37', 'E41', 'K2', 'E39', 'A39']
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")

    def _clear_pallet_list_sheet(self):
        """Очищает лист 'Список поддонов'"""
        if not hasattr(self, 'wb') or not self.wb:
            return False
            
        if "Список поддонов" not in self.wb.sheetnames:
            return False
            
        self.ws = self.wb["Список поддонов"]
        
        # Очищаем столбцы D, F, H, L (10-29)
        for row in range(10, 30):
            for col in ['D', 'F', 'H', 'L']:
                try:
                    cell = self.ws[f'{col}{row}']
                    cell.value = None
                    cell.alignment = Alignment(horizontal='general', vertical='center')
                except Exception as e:
                    print(f"Не удалось очистить ячейку {col}{row}: {e}")
        
        # Очищаем базовую информацию
        basic_info_cells = [
            'D7', 'D3', 'D6', 'D8', 'D37', 'E41', 'H3', 
            'D4', 'G4', 'E39', 'A39'
        ]
        
        for cell_address in basic_info_cells:
            try:
                cell = self.ws[cell_address]
                cell.value = None
                cell.alignment = Alignment(horizontal='general', vertical='center')
            except Exception as e:
                print(f"Не удалось очистить ячейку {cell_address}: {e}")
        
        return True
    
    def _get_tu_number(self):
        """Получает TU номер на основе выбранных manufacturer_var и product_type_var"""
        try:
            if not self.roll_module:
                return ""
                
            manufacturer = self.roll_module.manufacturer_var.get()
            product_type = self.roll_module.product_type_var.get()
            
            if not manufacturer or not product_type:
                return ""          
            
            # Ищем точное соответствие в packaging_tu.json
            packaging_data = self.config_manager.load_json_settings("packaging_tu.json")
            technical_specs = packaging_data.get("technical_specifications", [])
            
            for spec in technical_specs:
                if (spec["manufacturer"]["name"] == manufacturer and 
                    spec["product"]["name"] == product_type):
                    return spec["product"]["tu_number"]
                    
        except Exception as e:
            print(f"Ошибка получения TU номера: {e}")
        
        return "ТУ 9570-001-26604209-2014"  # Fallback
            
    def _is_file_locked(self, filepath):
        """Проверяет, открыт ли файл в другом процессе (например, в Excel)"""
        try:
            # Попытка открыть файл в режиме записи
            with open(filepath, 'a', encoding='utf-8'):
                pass
            return False
        except IOError:
            return True
        except Exception:
            return False
    
    def _set_cell_value(self, cell_address, value):
        """Устанавливает значение ячейки с обработкой ошибок"""
        try:
            # Проверяем, не является ли ячейка частью объединения
            cell = self.ws[cell_address]
            
            # Если ячейка объединенная - находим первую ячейку объединенного диапазона
            for merged_range in self.ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Для горизонтально объединенных ячеек (B11:E11, B12:E12 и т.д.)
                    # используем первую ячейку объединенного диапазона
                    target_cell = self.ws[merged_range.min_row][merged_range.min_col - 1]
                    target_cell.value = value
                    
                    # Выравнивание
                    if value is None or value == "":
                        target_cell.alignment = Alignment(horizontal='general', vertical='center')
                    elif isinstance(value, (int, float)):
                        target_cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        target_cell.alignment = Alignment(horizontal='left', vertical='center')
                    return
                        
            # Если ячейка не объединенная - работаем как обычно
            self.ws[cell_address] = value
            
            # Выравнивание
            if value is None or value == "":
                self.ws[cell_address].alignment = Alignment(horizontal='general', vertical='center')
            elif isinstance(value, (int, float)):
                self.ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.ws[cell_address].alignment = Alignment(horizontal='left', vertical='center')
                
        except Exception as e:
            print(f"Не удалось установить значение {value} в ячейку {cell_address}: {e}")
            
    def _update_manufacturer_info(self):
        """Обновляет информацию об изготовителе в первом листе файла для обоих цехов"""
        try:
            # Получаем значение чекбокса "Без изготовителя" из UI
            # True = чекбокс отмечен = "Без производителя" = НЕ показывать
            show_manufacturer = True

            if hasattr(self.roll_module, 'show_manufacturer_var'):
                # Инвертируем логику: True = "Без Производителя" = не показывать
                show_manufacturer = not self.roll_module.show_manufacturer_var.get()

            # Получаем выбранного производителя ИЗ UI (комбобокс)
            manufacturer_name = ""
            if hasattr(self.roll_module, 'manufacturer_var'):
                manufacturer_name = self.roll_module.manufacturer_var.get()
            
            # Формируем текст для отображения
            display_text = ""
            
            # Если чекбокс "Без изготовителя" НЕ отмечен (show_manufacturer=True) 
            # И есть выбранный производитель в UI
            if show_manufacturer and manufacturer_name:
                # Адрес для указанных производителей
                address = 'Россия, 426039, Удмуртская Республика, г. Ижевск, ул. Воткинское шоссе, д. 186, офис 1'
                
                if "Ремас" in manufacturer_name:
                    display_text = f'ООО "Ремас-Флексо", {address}'
                elif "Зюдин" in manufacturer_name:
                    display_text = f'ИП Зюдин В.Г., {address}'
                else:
                    # Для других изготовителей - только название из UI
                    display_text = manufacturer_name
            
            # Определяем, какой файл используется
            is_second_file = self._is_second_file()
            
            # Обновляем лист в зависимости от цеха
            if is_second_file:
                # Для 2 цеха: обновляем лист "Поддон" ячейка A1
                if "Поддон" in self.wb.sheetnames:
                    sheet = self.wb["Поддон"]
                    sheet['A1'] = display_text
                    
                    # Выравнивание по центру и перенос текста
                    sheet['A1'].alignment = Alignment(
                        horizontal='center', 
                        vertical='center',
                        wrap_text=True
                    )
                    print(f"Обновлен производитель для 2 цеха в ячейке A1: {display_text}")
            else:
                # Для 1 цеха: обновляем лист "Лист для коробки" ячейка B1
                if "Лист для коробки" in self.wb.sheetnames:
                    sheet = self.wb["Лист для коробки"]
                    sheet['B1'] = display_text
                    
                    # Выравнивание по центру и перенос текста
                    sheet['B1'].alignment = Alignment(
                        horizontal='center', 
                        vertical='center',
                        wrap_text=True
                    )
                    print(f"Обновлен производитель для 1 цеха в ячейке B1: {display_text}")
                
        except Exception as e:
            print(f"Ошибка при обновлении информации об изготовителе: {e}")
        
    def _convert_to_number_if_possible(self, value):
        """Пытается преобразовать строку в число"""
        if value is None:
            return None
            
        if not isinstance(value, str):
            return value
            
        # Убираем лишние пробелы
        value = value.strip()
        
        if not value:
            return None
        
        # Пробуем преобразовать в целое число
        try:
            if value.isdigit():
                return int(value)
        except:
            pass
        
        # Пробуем преобразовать в дробное число
        try:
            # Заменяем запятую на точку для дробных чисел
            normalized_value = value.replace(',', '.')
            # Проверяем, что это число с плавающей точкой
            if normalized_value.replace('.', '').isdigit() and normalized_value.count('.') == 1:
                return float(normalized_value)
        except:
            pass
        
        # Если не получилось - возвращаем как строку
        return value
    
    def on_settings_changed(self):
        """Обработчик изменений настроек от координатора"""
        # При изменении цеха можно обновить путь к файлу
        if hasattr(self, 'coordinator') and self.coordinator:
            workshop = self.coordinator.get_workshop()
            
    def _export_to_multitype_sheet_first_workshop(self, pallet_data):
        """Экспортирует данные в лист 'Много видов' с пересчетом с нуля"""
        try: 
            actual_file_path = self.get_excel_file_path()
            workbook = load_workbook(actual_file_path)
            pallet_sheet = workbook["Лист для паллеты"]
            
            boxes_count = 0
            gross_total = 0
            net_total = 0
            labels_total = 0
            
            for row in range(14, 29):
                if (pallet_sheet[f'B{row}'].value is not None or 
                    pallet_sheet[f'C{row}'].value is not None or
                    pallet_sheet[f'D{row}'].value is not None):
                    
                    boxes_count += 1
                    gross_total += pallet_sheet[f'B{row}'].value or 0
                    net_total += pallet_sheet[f'C{row}'].value or 0
                    labels_total += pallet_sheet[f'D{row}'].value or 0
            
            for row in range(14, 29):
                if (pallet_sheet[f'F{row}'].value is not None or 
                    pallet_sheet[f'G{row}'].value is not None or
                    pallet_sheet[f'H{row}'].value is not None):
                    
                    boxes_count += 1
                    gross_total += pallet_sheet[f'F{row}'].value or 0
                    net_total += pallet_sheet[f'G{row}'].value or 0
                    labels_total += pallet_sheet[f'H{row}'].value or 0
                        
            if boxes_count == 0:
                workbook.close()
                return {'success': False, 'error': 'Лист поддона 1 цеха пуст'}
            
            pallet_weight = self._convert_to_number_if_possible(pallet_data.get("pallet_weight", 0))
            gross_total_with_pallet = gross_total + pallet_weight
                
            multitype_sheet = workbook["Лист много видов"]
            
            product_name = pallet_data.get('product_name', '')
            target_row = None
            
            for row in range(11, 29):
                if multitype_sheet[f'B{row}'].value == product_name:
                    target_row = row
                    break
            
            if target_row is not None:
                multitype_sheet[f'A{target_row}'].value = None
                multitype_sheet[f'F{target_row}'].value = None
                multitype_sheet[f'G{target_row}'].value = None
                multitype_sheet[f'H{target_row}'].value = None
            
            if target_row is None:
                for row in range(11, 29):
                    if multitype_sheet[f'A{row}'].value is None:
                        target_row = row
                        break
            
            if target_row is None:
                workbook.close()
                return {'success': False, 'error': 'Лист переполнен'}
            
            # Временно сохраняем multitype_sheet как self.ws
            original_ws = getattr(self, 'ws', None)
            self.ws = multitype_sheet
            
            # Экспортируем базовую информацию без наименования продукции
            self._export_basic_info_for_first_workshop(skip_product_name=True)
            
            # Тип упаковки (поддон) и вес поддона
            self._set_cell_value('D6', pallet_data.get("pallet_type", ""))
            self._set_cell_value('K2', pallet_weight)
            
            # Записываем пересчитанные данные
            self._set_cell_value(f'A{target_row}', boxes_count)
            self._set_cell_value(f'B{target_row}', product_name)
            self._set_cell_value(f'F{target_row}', gross_total)
            self._set_cell_value(f'G{target_row}', net_total)
            self._set_cell_value(f'H{target_row}', labels_total)
            
            # Восстанавливаем рабочий лист
            if original_ws:
                self.ws = original_ws
            
            workbook.save(actual_file_path)
            workbook.close()
            
            return {'success': True, 'row_used': target_row}
            
        except Exception as e:
            print(f"Ошибка экспорта: {e}")
            try:
                workbook.close()
            except:
                pass
            return {'success': False, 'error': str(e)}          
            
    def _export_to_multitype_sheet_for_second_workshop(self, pallet_data):
        """Экспортирует данные в лист 'Много видов' для цеха 2"""      
        try: 
            actual_file_path = self.get_excel_file_path()           
            workbook = load_workbook(actual_file_path)           
            list_sheet = workbook["Список поддонов"]
            
            pallets_count = 0
            weight_total = 0
            quantity_total = 0
            length_total = 0
            
            for row in range(10, 30):
                if (list_sheet[f'F{row}'].value is not None or 
                    list_sheet[f'H{row}'].value is not None):
                    
                    pallets_count += 1
                    weight_total += list_sheet[f'F{row}'].value or 0
                    quantity_total += list_sheet[f'H{row}'].value or 0
                    
                    length = list_sheet[f'L{row}'].value
                    if length is not None:
                        length_total += length         
            
            if pallets_count == 0:
                workbook.close()
                return {'success': False, 'error': 'Лист поддона 2 цеха пуст'}
            
            multitype_sheet = workbook["Много видов"]
            
            product_name = pallet_data.get('product_name', '')         
            target_row = None
            
            for row in range(10, 30):
                if multitype_sheet[f'B{row}'].value == product_name:
                    target_row = row
                    break
            
            if target_row is not None:
                multitype_sheet[f'A{target_row}'].value = None
                multitype_sheet[f'H{target_row}'].value = None
                multitype_sheet[f'I{target_row}'].value = None
                multitype_sheet[f'L{target_row}'].value = None
            
            if target_row is None:
                for row in range(10, 30):
                    if multitype_sheet[f'A{row}'].value is None:
                        target_row = row
                        break
            
            if target_row is None:
                workbook.close()
                return {'success': False, 'error': 'Лист переполнен'}
            
            original_ws = getattr(self, 'ws', None)
            self.ws = multitype_sheet
            
            #  Вызов базовой информации
            self._export_basic_info_for_second_workshop(skip_product_name=True)
            self._export_box_data_for_second_workshop()
            
            pallet_weight = self._convert_to_number_if_possible(pallet_data.get("pallet_weight", 0))
            self._set_cell_value('H3', pallet_weight)
            
            # запись данных
            self._set_cell_value(f'A{target_row}', pallets_count)
            self._set_cell_value(f'B{target_row}', product_name)
            self._set_cell_value(f'H{target_row}', weight_total)
            self._set_cell_value(f'I{target_row}', quantity_total)
            self._set_cell_value(f'L{target_row}', length_total)
            
            # Восстанавливаем
            if original_ws:
                self.ws = original_ws
            
            workbook.save(actual_file_path)
            workbook.close()
            
            return {'success': True, 'row_used': target_row}
            
        except Exception as e:
            print(f"Ошибка экспорта в много-видовой лист для цеха 2: {e}")
            try:
                workbook.close()
            except:
                pass
            return {'success': False, 'error': str(e)}