# data_providers/workshop1_provider.py
from typing import Dict, Any
from openpyxl import load_workbook
import os
from .base_provider import BaseDataProvider

class Workshop1DataProvider(BaseDataProvider):
    """
    Специализированные методы для цеха 1
    """
    
    def __init__(self, roll_module, config_manager=None, excel_file_path=""):
        super().__init__(roll_module, config_manager, excel_file_path)

    def get_data_for_workshop1_box_noweight(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист 'ПоддонРолики' (коробка без веса).
        """
        all_data = self.collect_all_data()

        # В этом листе вес не нужен (data['has_weight'] = False)
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Количества
            'pallet_num': all_data['quantities'].get('pallet_num'),
            # Данные 'rolls_count' и 'quantity_per_roll' попадут в динамическую секцию.
            # Провайдер отдает их как есть, а маппинг уже знает, куда их вставлять.
            'rolls_count': all_data['quantities'].get('rolls_count'),
            'quantity_per_roll': all_data['quantities'].get('quantity_per_roll'),

            # Производитель
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),

            # Дополнительно
            'workshop': '1',
            'sheet_type': 'box_noweight',
            'has_weight': False  # Явно указываем, что веса нет
        }

    def get_data_for_workshop1_box(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист коробки.
        Возвращает только нужные для этого листа данные в удобном формате.
        """
        all_data = self.collect_all_data()
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'box_type': all_data['common'].get('box_type'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Веса
            'box_weight': all_data['weights'].get('box_weight'),
            'gross_weight_per_roll': all_data['weights'].get('gross_weight_per_roll'),
            'net_weight_per_roll': all_data['weights'].get('net_weight_per_roll'),
            
            # Количества
            'rolls_count': all_data['quantities'].get('rolls_count'),
            'quantity_per_roll': all_data['quantities'].get('quantity_per_roll'),
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'box',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }
    
    def get_data_for_workshop1_pallet(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист поддона.
        Включает данные о поддоне и коробках.
        """
        all_data = self.collect_all_data()
        
        # Получаем количество коробок
        boxes_count = all_data['quantities'].get('boxes_count', 0)
        if boxes_count == 0:
            boxes_count = 1
        
        return {
            # Основная информация (как для коробки)
            'customer': all_data['common'].get('customer'),            
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Данные поддона
            'pallet_type': all_data['common'].get('pallet_type'),
            'pallet_weight': all_data['weights'].get('pallet_weight'),
            'boxes_count': boxes_count,
            
            # Данные одной коробки (для заполнения в динамические секции)
            'gross_weight_per_box': all_data['weights'].get('total_gross'),
            'net_weight_per_box': all_data['weights'].get('total_net'),
            'quantity_per_box': all_data['quantities'].get('total_quantity'),
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'pallet',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }
        
    def get_data_for_workshop1_noweight(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист БезВеса (поддон без веса).
        Только количество, без весовых полей.
        """
        all_data = self.collect_all_data()
        
        # Получаем количество коробок
        boxes_count = all_data['quantities'].get('boxes_count', 0)
        if boxes_count == 0:
            boxes_count = 1
        
        # Получаем общее количество этикеток
        total_quantity = all_data['quantities'].get('total_quantity', 0)
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Данные для динамических секций
            'boxes_count': boxes_count,
            'quantity_per_box': total_quantity,  # Только количество
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'noweight',
            'has_weight': False  # Всегда False для этого метода
        }
    
    def get_data_for_workshop1_multitype(self) -> Dict[str, Any]:
        all_data = self.collect_all_data()
        excel_data = self._read_pallet_sheet_data()
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'order_number': all_data['common'].get('order_number'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            'product_text': all_data['common'].get('product_text'),  # Это же product_name
            
            # Данные из Excel
            'boxes_count': excel_data.get('boxes_count', 0),
            'gross_total': excel_data.get('gross_total', 0),
            'net_total': excel_data.get('net_total', 0),
            'labels_total': excel_data.get('labels_total', 0),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'multitype',
            'has_weight': True
        }

    def _read_pallet_sheet_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа поддона в Excel файле.
        Возвращает суммы по колонкам B, C, D (левая секция) и F, G, H (правая секция).
        """
        try:
            # Получаем путь к файлу ДЛЯ ЦЕХА 1 (фиксировано)
            actual_file_path = self.get_excel_file_path("1")  # Явно указываем цех 1

            if not os.path.exists(actual_file_path):
                return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

            workbook = load_workbook(actual_file_path, data_only=True)

            # Ищем нужный лист (может быть "Лист для паллеты" или другой)
            sheet_name = None
            possible_sheet_names = ["Лист для паллеты"]

            for name in possible_sheet_names:
                if name in workbook.sheetnames:
                    sheet_name = name
                    break

            if not sheet_name:
                workbook.close()
                return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

            pallet_sheet = workbook[sheet_name]

            boxes_count = 0
            gross_total = 0
            net_total = 0
            labels_total = 0

            # Левая секция: колонки B, C, D (строки 14-28)
            for row in range(14, 29):
                if any(pallet_sheet[f'{col}{row}'].value is not None
                       for col in ['B', 'C', 'D']):
                    boxes_count += 1
                    gross_total += pallet_sheet[f'B{row}'].value or 0
                    net_total += pallet_sheet[f'C{row}'].value or 0
                    labels_total += pallet_sheet[f'D{row}'].value or 0

            # Правая секция: колонки F, G, H (строки 14-28)
            for row in range(14, 29):
                if any(pallet_sheet[f'{col}{row}'].value is not None
                       for col in ['F', 'G', 'H']):
                    boxes_count += 1
                    gross_total += pallet_sheet[f'F{row}'].value or 0
                    net_total += pallet_sheet[f'G{row}'].value or 0
                    labels_total += pallet_sheet[f'H{row}'].value or 0

            workbook.close()

            return {
                'boxes_count': boxes_count,
                'gross_total': gross_total,
                'net_total': net_total,
                'labels_total': labels_total
            }

        except Exception as e:
            print(f"Ошибка чтения Excel файла: {e}")
            return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

    def get_data_for_workshop1_multitype_noweight(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист 'Много видов БезВеса'.
        Читает данные из листа 'БезВеса' для заполнения динамической секции.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'БезВеса'
        noweight_data = self._read_workshop1_noweight_sheet_data()

        return {
            # Основная информация (из UI)
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),
            'order_number': all_data['common'].get('order_number'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # product_text = product_name (из UI)
            'product_text': all_data['common'].get('product_text'),

            # Данные из листа 'БезВеса' (для динамической секции)
            'boxes_count': noweight_data.get('boxes_count', 0),
            'labels_total': noweight_data.get('labels_total', 0),

            # Производитель
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),

            # Дополнительно
            'workshop': '1',
            'sheet_type': 'multitype_noweight',
            'has_weight': False
        }

    def _read_workshop1_noweight_sheet_data(self) -> Dict[str, Any]:
        """Читает данные из листа 'БезВеса' Excel файла.
           Возвращает:
           - boxes_count: количество заполненных строк (для колонки A в Много видов)
           - labels_total: сумма количеств из колонок C, F, I (для колонки G в Много видов)
        """
        try:
            actual_file_path = self.get_excel_file_path("1")
            if not os.path.exists(actual_file_path):
                return {'boxes_count': 0, 'labels_total': 0}

            workbook = load_workbook(actual_file_path, data_only=True)

            if "БезВеса" not in workbook.sheetnames:
                workbook.close()
                return {'boxes_count': 0, 'labels_total': 0}

            sheet = workbook["БезВеса"]

            boxes_count = 0
            labels_total = 0

            # Строки 14-28 (как в маппинге noweight)
            for row in range(14, 29):
                row_filled = False

                # Левая секция (количество в C)
                qty_c = sheet[f'C{row}'].value
                if qty_c is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_c)
                    except:
                        pass

                # Центральная секция (количество в F)
                qty_f = sheet[f'F{row}'].value
                if qty_f is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_f)
                    except:
                        pass

                # Правая секция (количество в I)
                qty_i = sheet[f'I{row}'].value
                if qty_i is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_i)
                    except:
                        pass

                # Если хоть одна ячейка в строке заполнена - увеличиваем счётчик коробок
                if row_filled:
                    boxes_count += 1

            workbook.close()

            return {'boxes_count': boxes_count, 'labels_total': labels_total}

        except Exception as e:
            print(f"Ошибка чтения листа 'БезВеса': {e}")
            return {'boxes_count': 0, 'labels_total': 0}
    