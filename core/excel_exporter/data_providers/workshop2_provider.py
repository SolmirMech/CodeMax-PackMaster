# data_providers/workshop2_provider.py
from typing import Dict, Any
from openpyxl import load_workbook
import os
from .base_provider import BaseDataProvider

class Workshop2DataProvider(BaseDataProvider):
    """
    Специализированные методы для цеха 2
    """
    
    def __init__(self, roll_module, config_manager=None, excel_file_path="", coordinator=None):
        super().__init__(roll_module, config_manager, excel_file_path, coordinator)
    
    def get_data_for_workshop2_box(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Поддон' (коробка).
        Только нетто вес, длины роликов, данные о втулке.
        """
        all_data = self.collect_all_data()

        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),  # Тип упаковки (D3)
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Вес поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),  # Конвертировано из граммов
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные роликов
            'rolls_count': all_data['quantities'].get('rolls_count'),
            'pallet_num': all_data['quantities'].get('pallet_num'),
            'net_weight_per_roll': all_data['weights'].get('net_weight_per_roll'),  # Только нетто!
            'quantity_per_roll': all_data['quantities'].get('quantity_per_roll'),
            'roll_length': all_data['dimensions'].get('roll_length'),  # Длина ролика

            # Производитель
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'box',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }

    def get_data_for_workshop2_pallet_list(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Список поддонов'.
        Включает данные из листа 'Поддон' и рассчитывает итоги.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'Поддон' в Excel
        pallet_data = self._read_workshop2_pallet_sheet_data()

        # Берем данные производителя из основного кэша
        manufacturer_display_text = all_data['manufacturer'].get('display_text', '')

        return {
            # Основная информация (копируется из листа 'Поддон')
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Итоговые данные (рассчитываются из листа 'Поддон')
            # Для заполнения в динамические секции:
            'rolls_count': pallet_data.get('rolls_count', 0),  # Столбец D
            'total_weight': pallet_data.get('total_weight', 0),  # Столбец F
            'total_quantity': pallet_data.get('total_quantity', 0),  # Столбец H
            'total_length': pallet_data.get('total_length', 0),  # Столбец L

            # Производитель
            'manufacturer_display_text': manufacturer_display_text,

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'pallet_list',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }

    def _read_workshop2_pallet_sheet_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа 'Поддон' для 2 цеха и рассчитывает итоги.
        Аналогично WeightOrdersExporter._calculate_pallet_totals
        """
        try:
            # Получаем путь к файлу для 2 цеха
            actual_file_path = self.get_excel_file_path("2")

            if not os.path.exists(actual_file_path):
                return {
                    'rolls_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            workbook = load_workbook(actual_file_path, data_only=True)

            if "Поддон" not in workbook.sheetnames:
                workbook.close()
                return {
                    'rolls_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            pallet_sheet = workbook["Поддон"]

            total_quantity = 0
            total_weight = 0
            total_length = 0
            rolls_count = 0

            # Пары колонок и соответствующие смещения для длины в L
            column_pairs = [
                ('B', 'C', 0),  # B, C - длина в L с тем же номером строки
                ('E', 'F', 20),  # E, F - длина в L со смещением +20
                ('H', 'I', 40)  # H, I - длина в L со смещением +40
            ]

            for weight_col, qty_col, l_offset in column_pairs:
                for row in range(10, 30):  # строки 10-29
                    weight = pallet_sheet[f'{weight_col}{row}'].value
                    quantity = pallet_sheet[f'{qty_col}{row}'].value

                    if weight is not None or quantity is not None:
                        rolls_count += 1

                        if weight is not None:
                            total_weight += weight
                        if quantity is not None:
                            total_quantity += quantity

                        # Длина из столбца L
                        length_row = row + l_offset
                        length = pallet_sheet[f'L{length_row}'].value
                        if length is not None:
                            total_length += length

            workbook.close()

            return {
                'rolls_count': rolls_count,
                'total_weight': total_weight,
                'total_quantity': total_quantity,
                'total_length': total_length
            }

        except Exception as e:
            print(f"Ошибка чтения листа 'Поддон' для 2 цеха: {e}")
            return {
                'rolls_count': 0,
                'total_weight': 0,
                'total_quantity': 0,
                'total_length': 0
            }

    def get_data_for_workshop2_multitype(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Много видов'.
        Собирает данные из листа 'Список поддонов' для экспорта.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'Список поддонов' в Excel
        pallet_list_data = self._read_workshop2_pallet_list_data()

        # Получаем данные о производителе
        manufacturer_display_text = all_data['manufacturer'].get('display_text', '')

        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),  # Тип упаковки (D3)
            'order_number': all_data['common'].get('order_number'),
            # Примечание: product_text НЕ заполняется для этого режима (skip_product_name=True)
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Данные из листа 'Список поддонов' (для динамических секций)
            'pallets_count': pallet_list_data.get('pallets_count', 0),  # Столбец A (количество поддонов)
            'product_name': all_data['common'].get('product_text', ''),  # Наименование из UI (столбец B)
            'total_weight': pallet_list_data.get('total_weight', 0),  # Столбец H (суммарный вес)
            'total_quantity': pallet_list_data.get('total_quantity', 0),  # Столбец I (суммарное количество)
            'total_length': pallet_list_data.get('total_length', 0),  # Столбец L (суммарная длина)

            # Производитель
            'manufacturer_display_text': manufacturer_display_text,

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'multitype',
            'has_weight': all_data['metadata'].get('has_weight', True)  # Всегда true для этого режима
        }

    def _read_workshop2_pallet_list_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа 'Список поддонов' для 2 цеха.
        Возвращает суммированные данные по листу.
        """
        try:
            # Получаем путь к файлу для 2 цеха
            actual_file_path = self.get_excel_file_path("2")

            if not os.path.exists(actual_file_path):
                return {
                    'pallets_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            workbook = load_workbook(actual_file_path, data_only=True)

            if "Список поддонов" not in workbook.sheetnames:
                workbook.close()
                return {
                    'pallets_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            list_sheet = workbook["Список поддонов"]

            # Суммируем данные из столбцов D, F, H, L (строки 10-29)
            total_weight = 0
            total_quantity = 0
            total_length = 0
            pallets_count = 0

            for row in range(10, 30):  # строки 10-29
                # Проверяем, есть ли данные в строке
                if (list_sheet[f'D{row}'].value is not None or
                        list_sheet[f'F{row}'].value is not None or
                        list_sheet[f'H{row}'].value is not None or
                        list_sheet[f'L{row}'].value is not None):

                    pallets_count += 1

                    # Суммируем значения из столбцов
                    if list_sheet[f'F{row}'].value is not None:
                        total_weight += list_sheet[f'F{row}'].value

                    if list_sheet[f'H{row}'].value is not None:
                        total_quantity += list_sheet[f'H{row}'].value

                    if list_sheet[f'L{row}'].value is not None:
                        total_length += list_sheet[f'L{row}'].value

            workbook.close()

            return {
                'pallets_count': pallets_count,
                'total_weight': total_weight,
                'total_quantity': total_quantity,
                'total_length': total_length
            }

        except Exception as e:
            print(f"Ошибка чтения листа 'Список поддонов' для 2 цеха: {e}")
            return {
                'pallets_count': 0,
                'total_weight': 0,
                'total_quantity': 0,
                'total_length': 0
            }
    