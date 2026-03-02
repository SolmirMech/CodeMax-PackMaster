# data_provider.py
import os
from core.config_manager import ConfigManager
from typing import Dict, Any, Optional, Union
from openpyxl import load_workbook


# noinspection SpellCheckingInspection
class ExportDataProvider:
    """
    Единый централизованный сборщик данных из UI.
    Не знает о структуре Excel, только о данных.
    """
    
    def __init__(self, roll_module, config_manager: Optional[ConfigManager] = None, excel_file_path: str = ""):
        """
        Инициализация сборщика данных.
        
        Args:
            roll_module: Модуль UI с данными (должен иметь стандартные атрибуты)
            config_manager: Менеджер конфигурации для TU номеров и настроек
        """
        self.roll_module = roll_module
        self.config_manager = config_manager or ConfigManager()
        self.original_excel_path = excel_file_path
        self._data_cache: Optional[Dict[str, Any]] = None
        
    def collect_all_data(self) -> Dict[str, Any]:
        """
        Собирает полный набор данных из UI для экспорта.
        Данные кешируются для повторного использования.
        
        Returns:
            Словарь с полными данными, структурированный по категориям
        """
        if self._data_cache is not None:
            return self._data_cache.copy()
            
        self._data_cache = {
            'common': self._get_common_data(),
            'weights': self._get_weight_data(),
            'quantities': self._get_quantity_data(),
            'dimensions': self._get_dimension_data(),
            'metadata': self._get_metadata(),
            'manufacturer': self._get_manufacturer_data()
        }
        
        return self._data_cache.copy()
    
    def clear_cache(self):
        """Очищает кеш данных (например, при изменении данных в UI)"""
        self._data_cache = None

    def get_data_for_workshop1_multitype_noweight(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист 'Много видов БезВеса'.
        Читает данные из листа 'БезВеса' для заполнения динамической секции.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'БезВеса'
        noweight_data = self._read_workshop1_noweight_sheet_data()

        return {
            # Основная информация (из UI)
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),
            'order_number': all_data['common'].get('order_number'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # product_text = product_name (из UI)
            'product_text': all_data['common'].get('product_text'),

            # Данные из листа 'БезВеса' (для динамической секции)
            'boxes_count': noweight_data.get('boxes_count', 0),
            'labels_total': noweight_data.get('labels_total', 0),

            # Производитель
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),

            # Дополнительно
            'workshop': '1',
            'sheet_type': 'multitype_noweight',
            'has_weight': False
        }

    def _read_workshop1_noweight_sheet_data(self) -> Dict[str, Any]:
        """Читает данные из листа 'БезВеса' Excel файла.
           Возвращает:
           - boxes_count: количество заполненных строк (для колонки A в Много видов)
           - labels_total: сумма количеств из колонок C, F, I (для колонки G в Много видов)
        """
        try:
            actual_file_path = self.get_excel_file_path("1")
            if not os.path.exists(actual_file_path):
                return {'boxes_count': 0, 'labels_total': 0}

            workbook = load_workbook(actual_file_path, data_only=True)

            if "БезВеса" not in workbook.sheetnames:
                workbook.close()
                return {'boxes_count': 0, 'labels_total': 0}

            sheet = workbook["БезВеса"]

            boxes_count = 0
            labels_total = 0

            # Строки 14-28 (как в маппинге noweight)
            for row in range(14, 29):
                row_filled = False

                # Левая секция (количество в C)
                qty_c = sheet[f'C{row}'].value
                if qty_c is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_c)
                    except:
                        pass

                # Центральная секция (количество в F)
                qty_f = sheet[f'F{row}'].value
                if qty_f is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_f)
                    except:
                        pass

                # Правая секция (количество в I)
                qty_i = sheet[f'I{row}'].value
                if qty_i is not None:
                    row_filled = True
                    try:
                        labels_total += float(qty_i)
                    except:
                        pass

                # Если хоть одна ячейка в строке заполнена - увеличиваем счётчик коробок
                if row_filled:
                    boxes_count += 1

            workbook.close()

            return {'boxes_count': boxes_count, 'labels_total': labels_total}

        except Exception as e:
            print(f"Ошибка чтения листа 'БезВеса': {e}")
            return {'boxes_count': 0, 'labels_total': 0}

    def get_data_for_workshop1_box(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист коробки.
        Возвращает только нужные для этого листа данные в удобном формате.
        """
        all_data = self.collect_all_data()
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'box_type': all_data['common'].get('box_type'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Веса
            'box_weight': all_data['weights'].get('box_weight'),
            'gross_weight_per_roll': all_data['weights'].get('gross_weight_per_roll'),
            'net_weight_per_roll': all_data['weights'].get('net_weight_per_roll'),
            
            # Количества
            'rolls_count': all_data['quantities'].get('rolls_count'),
            'quantity_per_roll': all_data['quantities'].get('quantity_per_roll'),
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'box',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }
    
    def get_data_for_workshop1_pallet(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист поддона.
        Включает данные о поддоне и коробках.
        """
        all_data = self.collect_all_data()
        
        # Получаем количество коробок
        boxes_count = all_data['quantities'].get('boxes_count', 0)
        if boxes_count == 0:
            boxes_count = 1
        
        return {
            # Основная информация (как для коробки)
            'customer': all_data['common'].get('customer'),            
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Данные поддона
            'pallet_type': all_data['common'].get('pallet_type'),
            'pallet_weight': all_data['weights'].get('pallet_weight'),
            'boxes_count': boxes_count,
            
            # Данные одной коробки (для заполнения в динамические секции)
            'gross_weight_per_box': all_data['weights'].get('total_gross'),
            'net_weight_per_box': all_data['weights'].get('total_net'),
            'quantity_per_box': all_data['quantities'].get('total_quantity'),
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'pallet',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }
        
    def get_data_for_workshop1_noweight(self) -> Dict[str, Any]:
        """
        Специализированный метод для 1 цеха, лист БезВеса (поддон без веса).
        Только количество, без весовых полей.
        """
        all_data = self.collect_all_data()
        
        # Получаем количество коробок
        boxes_count = all_data['quantities'].get('boxes_count', 0)
        if boxes_count == 0:
            boxes_count = 1
        
        # Получаем общее количество этикеток
        total_quantity = all_data['quantities'].get('total_quantity', 0)
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            
            # Данные для динамических секций
            'boxes_count': boxes_count,
            'quantity_per_box': total_quantity,  # Только количество
            
            # Производитель
            'show_manufacturer': all_data['manufacturer'].get('show_manufacturer'),
            'manufacturer_name': all_data['manufacturer'].get('manufacturer_name'),
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'noweight',
            'has_weight': False  # Всегда False для этого метода
        }
    
    def get_data_for_workshop1_multitype(self) -> Dict[str, Any]:
        all_data = self.collect_all_data()
        excel_data = self._read_pallet_sheet_data()
        
        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'order_number': all_data['common'].get('order_number'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),
            'product_text': all_data['common'].get('product_text'),  # Это же product_name
            
            # Данные из Excel
            'boxes_count': excel_data.get('boxes_count', 0),
            'gross_total': excel_data.get('gross_total', 0),
            'net_total': excel_data.get('net_total', 0),
            'labels_total': excel_data.get('labels_total', 0),
            
            # Дополнительно
            'workshop': '1',
            'sheet_type': 'multitype',
            'has_weight': True
        }

    def get_data_for_workshop2_box(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Поддон' (коробка).
        Только нетто вес, длины роликов, данные о втулке.
        """
        all_data = self.collect_all_data()

        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),  # Тип упаковки (D3)
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Вес поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),  # Конвертировано из граммов
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные роликов
            'rolls_count': all_data['quantities'].get('rolls_count'),
            'pallet_num': all_data['quantities'].get('pallet_num'),
            'net_weight_per_roll': all_data['weights'].get('net_weight_per_roll'),  # Только нетто!
            'quantity_per_roll': all_data['quantities'].get('quantity_per_roll'),
            'roll_length': all_data['dimensions'].get('roll_length'),  # Длина ролика

            # Производитель
            'manufacturer_display_text': all_data['manufacturer'].get('display_text'),

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'box',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }

    def get_data_for_workshop2_pallet_list(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Список поддонов'.
        Включает данные из листа 'Поддон' и рассчитывает итоги.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'Поддон' в Excel
        pallet_data = self._read_workshop2_pallet_sheet_data()

        # Берем данные производителя из основного кэша
        manufacturer_display_text = all_data['manufacturer'].get('display_text', '')

        return {
            # Основная информация (копируется из листа 'Поддон')
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),
            'order_number': all_data['common'].get('order_number'),
            'product_text': all_data['common'].get('product_text'),
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Итоговые данные (расчитываются из листа 'Поддон')
            # Для заполнения в динамические секции:
            'rolls_count': pallet_data.get('rolls_count', 0),  # Столбец D
            'total_weight': pallet_data.get('total_weight', 0),  # Столбец F
            'total_quantity': pallet_data.get('total_quantity', 0),  # Столбец H
            'total_length': pallet_data.get('total_length', 0),  # Столбец L

            # Производитель
            'manufacturer_display_text': manufacturer_display_text,

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'pallet_list',
            'has_weight': all_data['metadata'].get('has_weight', False)
        }

    def _read_pallet_sheet_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа поддона в Excel файле.
        Возвращает суммы по колонкам B, C, D (левая секция) и F, G, H (правая секция).
        """
        try:
            # Получаем путь к файлу ДЛЯ ЦЕХА 1 (фиксировано)
            actual_file_path = self.get_excel_file_path("1")  # Явно указываем цех 1

            if not os.path.exists(actual_file_path):
                return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

            workbook = load_workbook(actual_file_path, data_only=True)

            # Ищем нужный лист (может быть "Лист для паллеты" или другой)
            sheet_name = None
            possible_sheet_names = ["Лист для паллеты"]

            for name in possible_sheet_names:
                if name in workbook.sheetnames:
                    sheet_name = name
                    break

            if not sheet_name:
                workbook.close()
                return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

            pallet_sheet = workbook[sheet_name]

            boxes_count = 0
            gross_total = 0
            net_total = 0
            labels_total = 0

            # Левая секция: колонки B, C, D (строки 14-28)
            for row in range(14, 29):
                if any(pallet_sheet[f'{col}{row}'].value is not None
                       for col in ['B', 'C', 'D']):
                    boxes_count += 1
                    gross_total += pallet_sheet[f'B{row}'].value or 0
                    net_total += pallet_sheet[f'C{row}'].value or 0
                    labels_total += pallet_sheet[f'D{row}'].value or 0

            # Правая секция: колонки F, G, H (строки 14-28)
            for row in range(14, 29):
                if any(pallet_sheet[f'{col}{row}'].value is not None
                       for col in ['F', 'G', 'H']):
                    boxes_count += 1
                    gross_total += pallet_sheet[f'F{row}'].value or 0
                    net_total += pallet_sheet[f'G{row}'].value or 0
                    labels_total += pallet_sheet[f'H{row}'].value or 0

            workbook.close()

            return {
                'boxes_count': boxes_count,
                'gross_total': gross_total,
                'net_total': net_total,
                'labels_total': labels_total
            }

        except Exception as e:
            print(f"Ошибка чтения Excel файла: {e}")
            return {'boxes_count': 0, 'gross_total': 0, 'net_total': 0, 'labels_total': 0}

    def _read_workshop2_pallet_sheet_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа 'Поддон' для 2 цеха и рассчитывает итоги.
        Аналогично WeightOrdersExporter._calculate_pallet_totals
        """
        try:
            # Получаем путь к файлу для 2 цеха
            actual_file_path = self.get_excel_file_path("2")

            if not os.path.exists(actual_file_path):
                return {
                    'rolls_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            workbook = load_workbook(actual_file_path, data_only=True)

            if "Поддон" not in workbook.sheetnames:
                workbook.close()
                return {
                    'rolls_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            pallet_sheet = workbook["Поддон"]

            total_quantity = 0
            total_weight = 0
            total_length = 0
            rolls_count = 0

            # Пары колонок и соответствующие смещения для длины в L
            column_pairs = [
                ('B', 'C', 0),  # B, C - длина в L с тем же номером строки
                ('E', 'F', 20),  # E, F - длина в L со смещением +20
                ('H', 'I', 40)  # H, I - длина в L со смещением +40
            ]

            for weight_col, qty_col, l_offset in column_pairs:
                for row in range(10, 30):  # строки 10-29
                    weight = pallet_sheet[f'{weight_col}{row}'].value
                    quantity = pallet_sheet[f'{qty_col}{row}'].value

                    if weight is not None or quantity is not None:
                        rolls_count += 1

                        if weight is not None:
                            total_weight += weight
                        if quantity is not None:
                            total_quantity += quantity

                        # Длина из столбца L
                        length_row = row + l_offset
                        length = pallet_sheet[f'L{length_row}'].value
                        if length is not None:
                            total_length += length

            workbook.close()

            return {
                'rolls_count': rolls_count,
                'total_weight': total_weight,
                'total_quantity': total_quantity,
                'total_length': total_length
            }

        except Exception as e:
            print(f"Ошибка чтения листа 'Поддон' для 2 цеха: {e}")
            return {
                'rolls_count': 0,
                'total_weight': 0,
                'total_quantity': 0,
                'total_length': 0
            }

    def get_data_for_workshop2_multitype(self) -> Dict[str, Any]:
        """
        Специализированный метод для 2 цеха, лист 'Много видов'.
        Собирает данные из листа 'Список поддонов' для экспорта.
        """
        all_data = self.collect_all_data()

        # Читаем данные из листа 'Список поддонов' в Excel
        pallet_list_data = self._read_workshop2_pallet_list_data()

        # Получаем данные о производителе
        manufacturer_display_text = all_data['manufacturer'].get('display_text', '')

        return {
            # Основная информация
            'customer': all_data['common'].get('customer'),
            'pallet_type': all_data['common'].get('pallet_type'),  # Тип упаковки (D3)
            'order_number': all_data['common'].get('order_number'),
            # Примечание: product_text НЕ заполняется для этого режима (skip_product_name=True)
            'date': all_data['common'].get('date'),
            'packer': all_data['common'].get('packer'),
            'product_type': all_data['common'].get('product_type'),
            'tu_number': all_data['common'].get('tu_number'),

            # Данные втулки
            'sleeve_weight_kg': all_data['weights'].get('sleeve_weight_kg'),
            'sleeve_diameter': all_data['dimensions'].get('sleeve_diameter'),

            # Данные поддона
            'pallet_weight': all_data['weights'].get('pallet_weight'),

            # Данные из листа 'Список поддонов' (для динамических секций)
            'pallets_count': pallet_list_data.get('pallets_count', 0),  # Столбец A (количество поддонов)
            'product_name': all_data['common'].get('product_text', ''),  # Наименование из UI (столбец B)
            'total_weight': pallet_list_data.get('total_weight', 0),  # Столбец H (суммарный вес)
            'total_quantity': pallet_list_data.get('total_quantity', 0),  # Столбец I (суммарное количество)
            'total_length': pallet_list_data.get('total_length', 0),  # Столбец L (суммарная длина)

            # Производитель
            'manufacturer_display_text': manufacturer_display_text,

            # Дополнительно
            'workshop': '2',
            'sheet_type': 'multitype',
            'has_weight': all_data['metadata'].get('has_weight', True)  # Всегда true для этого режима
        }

    def _read_workshop2_pallet_list_data(self) -> Dict[str, Any]:
        """
        Читает данные из листа 'Список поддонов' для 2 цеха.
        Возвращает суммированные данные по листу.
        """
        try:
            # Получаем путь к файлу для 2 цеха
            actual_file_path = self.get_excel_file_path("2")

            if not os.path.exists(actual_file_path):
                return {
                    'pallets_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            workbook = load_workbook(actual_file_path, data_only=True)

            if "Список поддонов" not in workbook.sheetnames:
                workbook.close()
                return {
                    'pallets_count': 0,
                    'total_weight': 0,
                    'total_quantity': 0,
                    'total_length': 0
                }

            list_sheet = workbook["Список поддонов"]

            # Суммируем данные из столбцов D, F, H, L (строки 10-29)
            total_weight = 0
            total_quantity = 0
            total_length = 0
            pallets_count = 0

            for row in range(10, 30):  # строки 10-29
                # Проверяем, есть ли данные в строке
                if (list_sheet[f'D{row}'].value is not None or
                        list_sheet[f'F{row}'].value is not None or
                        list_sheet[f'H{row}'].value is not None or
                        list_sheet[f'L{row}'].value is not None):

                    pallets_count += 1

                    # Суммируем значения из столбцов
                    if list_sheet[f'F{row}'].value is not None:
                        total_weight += list_sheet[f'F{row}'].value

                    if list_sheet[f'H{row}'].value is not None:
                        total_quantity += list_sheet[f'H{row}'].value

                    if list_sheet[f'L{row}'].value is not None:
                        total_length += list_sheet[f'L{row}'].value

            workbook.close()

            return {
                'pallets_count': pallets_count,
                'total_weight': total_weight,
                'total_quantity': total_quantity,
                'total_length': total_length
            }

        except Exception as e:
            print(f"Ошибка чтения листа 'Список поддонов' для 2 цеха: {e}")
            return {
                'pallets_count': 0,
                'total_weight': 0,
                'total_quantity': 0,
                'total_length': 0
            }

    # ==================== ПРИВАТНЫЕ МЕТОДЫ СБОРА ====================
    
    def _get_common_data(self) -> Dict[str, Any]:
        """Собирает общие данные, общие для всех листов"""
        if not self.roll_module:
            return {}
            
        data = {}
        
        try:
            # Заказчик
            if hasattr(self.roll_module, 'customer_var'):
                data['customer'] = self.roll_module.customer_var.get()
            
            # Название коробки
            if hasattr(self.roll_module, 'box_size_var'):
                data['box_type'] = self.roll_module.box_size_var.get()
                
            # Название паллеты
            data['pallet_type'] = self.roll_module.preview_module.export_module.pallet_size_var.get()
                         
            # Полный номер заказа
            if hasattr(self.roll_module, 'order_prefix'):
                order_prefix = self.roll_module.order_prefix.get()
                order_number = getattr(self.roll_module, 'order_number', '').get()
                order_suffix = getattr(self.roll_module, 'order_suffix', '').get()
                data['order_number'] = f"{order_prefix}{order_number}{order_suffix}"
            
            # Наименование продукции (многострочное)
            if hasattr(self.roll_module, 'product_text'):
                product_text = self.roll_module.product_text.get("1.0", "end-1c").strip()
                data['product_text'] = product_text
            
            # Дата упаковки
            if hasattr(self.roll_module, 'date_var'):
                data['date'] = self.roll_module.date_var.get()
            
            # Упаковщик
            if hasattr(self.roll_module, 'packer_var'):
                data['packer'] = self.roll_module.packer_var.get()
            
            # ВАЖНО: Используем готовые данные из модуля производителя
            if hasattr(self.roll_module, 'get_manufacturer_full_data'):
                manufacturer_data = self.roll_module.get_manufacturer_full_data()
                
                if manufacturer_data.get('tu_number') and manufacturer_data['tu_number'].strip() not in ["—", "-", ""]:
                    # Есть TU номер из модуля производителя
                    data['tu_number'] = manufacturer_data['tu_number']
                    data['product_type'] = manufacturer_data.get('product', '')
                elif hasattr(self.roll_module, 'product_type_var') and self.roll_module.product_type_var.get():
                    # Нет TU из XML, но есть выбор в комбобоксе
                    data['product_type'] = self.roll_module.product_type_var.get()
                    data['tu_number'] = self._get_tu_number()
                else:
                    # Ничего нет
                    data['product_type'] = ""
                    data['tu_number'] = "ТУ технические условия"
            else:
                # Fallback на старую логику
                if hasattr(self.roll_module, 'product_type_var'):
                    data['product_type'] = self.roll_module.product_type_var.get()
                    data['tu_number'] = self._get_tu_number()
                else:
                    data['product_type'] = ""
                    data['tu_number'] = "ТУ технические условия"
                
        except Exception as e:
            print(f"Ошибка сбора общих данных: {e}")
            # Возвращаем частично собранные данные
            pass
            
        return data
        
    def _get_tu_number(self) -> str:
        """Получает TU номер на основе производителя и типа продукта"""
        try:
            if not self.roll_module:
                return "ТУ технические условия"  # Fallback
                
            manufacturer = self.roll_module.manufacturer_var.get() if hasattr(self.roll_module, 'manufacturer_var') else ""
            
            # Используем уже вычисленный product_type из data, если есть
            product_type = self.roll_module.product_type_var.get() if hasattr(self.roll_module, 'product_type_var') else ""
            
            if not manufacturer or not product_type:
                return "ТУ технические условия"
            
            # Ищем в конфигурации
            packaging_data = self.config_manager.load_json_settings("packaging_tu.json")
            technical_specs = packaging_data.get("technical_specifications", [])
            
            for spec in technical_specs:
                if (spec["manufacturer"]["name"] == manufacturer and 
                    spec["product"]["name"] == product_type):
                    return spec["product"]["tu_number"]
                    
        except Exception as e:
            print(f"Ошибка получения ТУ номера: {e}")
        
        return "ТУ технические условия"  # Fallback
        
    def _get_product_type_by_tu(self, tu_number: str) -> str:
        """Получает тип продукта по TU номеру из конфигурации"""
        try:
            if not tu_number or not self.roll_module:
                return self.roll_module.product_type_var.get() if hasattr(self.roll_module, 'product_type_var') else ""
            
            # Ищем в конфигурации
            packaging_data = self.config_manager.load_json_settings("packaging_tu.json")
            technical_specs = packaging_data.get("technical_specifications", [])
            
            for spec in technical_specs:
                if spec["product"]["tu_number"] == tu_number:
                    return spec["product"]["name"]
                    
        except Exception as e:
            print(f"Ошибка получения типа продукта по ТУ номеру: {e}")
        
        # Fallback - возвращаем текущее значение из UI
        return self.roll_module.product_type_var.get() if hasattr(self.roll_module, 'product_type_var') else ""
    
    def _get_weight_data(self) -> Dict[str, Any]:
        """Собирает все данные, связанные с весом"""
        if not self.roll_module:
            return {}
            
        data = {}
        
        try:
            # Вес коробки
            if hasattr(self.roll_module, 'box_weight_var'):
                weight_str = self.roll_module.box_weight_var.get()
                data['box_weight'] = self._convert_to_number(weight_str)
                
            # Вес паллеты
            weight_str = self.roll_module.preview_module.export_module.pallet_weight_var.get()
            data['pallet_weight'] = self._convert_to_number(weight_str)
            
            # Общий вес брутто
            if hasattr(self.roll_module, 'total_gross_var'):
                weight_str = self.roll_module.total_gross_var.get()
                data['total_gross'] = self._convert_to_number(weight_str)
            
            # Общий вес нетто
            if hasattr(self.roll_module, 'total_net_var'):
                weight_str = self.roll_module.total_net_var.get()
                data['total_net'] = self._convert_to_number(weight_str)
            
            # Вес брутто одного ролика (кг)
            if hasattr(self.roll_module, 'gross_weight_kg_var'):
                weight_str = self.roll_module.gross_weight_kg_var.get()
                data['gross_weight_per_roll'] = self._convert_to_number(weight_str)
            
            # Вес нетто одного ролика (кг)
            if hasattr(self.roll_module, 'net_weight_kg_var'):
                weight_str = self.roll_module.net_weight_kg_var.get()
                data['net_weight_per_roll'] = self._convert_to_number(weight_str)
            
            # Вес втулки (граммы -> преобразуем в кг при необходимости)
            if hasattr(self.roll_module, 'sleeve_weight_var'):
                weight_str = self.roll_module.sleeve_weight_var.get()
                data['sleeve_weight_g'] = self._convert_to_number(weight_str)
                if data['sleeve_weight_g'] is not None:
                    data['sleeve_weight_kg'] = data['sleeve_weight_g'] / 1000
            
        except Exception as e:
            print(f"Ошибка сбора данных веса: {e}")
            
        return data
    
    def _get_quantity_data(self) -> Dict[str, Any]:
        """Собирает данные по количеству"""
        if not self.roll_module:
            return {}
            
        data = {}
        
        try:
            
            boxes_count_str = self.roll_module.preview_module.export_module.boxes_count_var.get()
            data['boxes_count'] = self._convert_to_number(boxes_count_str, force_int=True)
            
            pallet_num_str = self.roll_module.preview_module.export_module.pallet_num_var.get()
            data['pallet_num'] = self._convert_to_number(pallet_num_str, force_int=True)            
            
            # Количество роликов
            if hasattr(self.roll_module, 'rolls_count_var'):
                count_str = self.roll_module.rolls_count_var.get()
                data['rolls_count'] = self._convert_to_number(count_str, force_int=True)
            
            # Количество этикеток в ролике
            if hasattr(self.roll_module, 'quantity_var'):
                qty_str = self.roll_module.quantity_var.get()
                data['quantity_per_roll'] = self._convert_to_number(qty_str, force_int=True)
            
            # Общее количество этикеток
            if hasattr(self.roll_module, 'total_quantity_var'):
                qty_str = self.roll_module.total_quantity_var.get()
                data['total_quantity'] = self._convert_to_number(qty_str, force_int=True)
            
        except Exception as e:
            print(f"Ошибка сбора данных количества: {e}")
            
        return data

    def _get_dimension_data(self) -> Dict[str, Any]:
        """Собирает данные по размерам"""
        if not self.roll_module:
            return {}
            
        data = {}
        
        try:
            # Диаметр втулки
            if hasattr(self.roll_module, 'sleeve_diameter_var'):
                diam_str = self.roll_module.sleeve_diameter_var.get()
                data['sleeve_diameter'] = self._convert_to_number(diam_str)
            
            # Длина ролика
            if hasattr(self.roll_module, 'roll_length'):
                length_str = self.roll_module.roll_length.get()
                data['roll_length'] = self._convert_to_number(length_str)
            
        except Exception as e:
            print(f"Ошибка сбора данных размеров: {e}")
            
        return data
    
    def _get_metadata(self) -> Dict[str, Any]:
        """Собирает метаданные и состояние"""
        data = {}
        
        try:
            # Есть ли вес (определяем по total_gross)
            if hasattr(self.roll_module, 'total_gross_var'):
                weight_value = self.roll_module.total_gross_var.get()
                has_weight = bool(weight_value and str(weight_value).strip() and 
                                str(weight_value).strip() != '0')
                data['has_weight'] = has_weight
            else:
                data['has_weight'] = False
                
        except Exception as e:
            print(f"Ошибка сбора метаданных: {e}")
            data['has_weight'] = False
            
        return data            
    
    def _get_manufacturer_data(self) -> Dict[str, Any]:
        """Собирает данные по производителю"""
        data = {
            'show_manufacturer': True,  # По умолчанию показываем
            'manufacturer_name': '',
            'display_text': ''
        }
        
        if not self.roll_module:
            return data
            
        try:
            # Чекбокс "Без изготовителя" (True = отмечен = не показывать)
            if hasattr(self.roll_module, 'show_manufacturer_var'):
                # Инвертируем: True в UI = "Без производителя" = не показывать
                data['show_manufacturer'] = not self.roll_module.show_manufacturer_var.get()
            
            # Имя производителя из комбобокса
            if hasattr(self.roll_module, 'manufacturer_var'):
                data['manufacturer_name'] = self.roll_module.manufacturer_var.get()
            
            # Формируем текст для отображения с адресом из конфигурации
            if data['show_manufacturer'] and data['manufacturer_name']:
                # Ищем производителя в конфигурации для получения адреса
                manufacturer_address = self._get_manufacturer_address(data['manufacturer_name'])
                
                if manufacturer_address:
                    data['display_text'] = f'{data["manufacturer_name"]}, {manufacturer_address}'
                else:
                    # Если адрес не найден, показываем только имя
                    data['display_text'] = data['manufacturer_name']
                        
        except Exception as e:
            print(f"Ошибка сбора данных производителя: {e}")
            
        return data

    def _get_manufacturer_address(self, manufacturer_name: str) -> Optional[str]:
        """Получает адрес производителя из конфигурации packaging_tu.json"""
        try:
            if not manufacturer_name:
                return None
            
            # Ищем в конфигурации
            packaging_data = self.config_manager.load_json_settings("packaging_tu.json")
            technical_specs = packaging_data.get("technical_specifications", [])
            
            # Ищем точное совпадение по имени производителя
            for spec in technical_specs:
                if spec["manufacturer"]["name"] == manufacturer_name:
                    return spec["manufacturer"]["address"]
            
            # Если не нашли точного совпадения, пробуем частичное (на всякий случай)
            for spec in technical_specs:
                if manufacturer_name in spec["manufacturer"]["name"] or spec["manufacturer"]["name"] in manufacturer_name:
                    return spec["manufacturer"]["address"]
                    
        except Exception as e:
            print(f"Ошибка получения адреса производителя: {e}")
        
        return None
    
    @staticmethod
    def _convert_to_number(value: Optional[str], force_int: bool = False) -> Optional[Union[int, float]]:
        """
        Безопасно преобразует строку в число.
        
        Args:
            value: Строка для преобразования
            force_int: Принудительно возвращать целое число
            
        Returns:
            Число (int или float) или None
        """
        if value is None:
            return None
            
        if not isinstance(value, str):
            # Если уже число - возвращаем как есть
            if isinstance(value, (int, float)):
                return int(value) if force_int and value.is_integer() else value
            return None
        
        value = value.strip()
        if not value:
            return None
        
        # Пробуем целое число
        try:
            if value.isdigit():
                return int(value)
        except:
            pass
        
        # Пробуем дробное число
        try:
            normalized = value.replace(',', '.')
            # Проверяем формат числа с плавающей точкой
            if normalized.replace('.', '').isdigit() and normalized.count('.') == 1:
                result = float(normalized)
                return int(result) if force_int and result.is_integer() else result
        except:
            pass
        
        return None
        
    def get_excel_file_path(self, workshop: str) -> str:
        """Определяет путь к файлу Excel"""
        # Копируем логику из старого get_excel_file_path
        try:
            # Получаем путь из настроек
            settings = self.config_manager.load_json_settings("shared_utils.json")
            excel_folder = settings.get("weight_orders_xlsx", "")
            
            if not excel_folder:
                excel_folder = os.path.dirname(self.original_excel_path)
            
            # Определяем имя файла
            filename = "weight_orders.xlsx" if workshop == "1" else "weight_orders_2.xlsx"
            full_path = os.path.join(excel_folder, filename)
            
            # Проверяем существование файла
            if not os.path.exists(full_path):
                # Копируем из assets
                assets_file = self.config_manager.get_asset_path(filename)
                if os.path.exists(assets_file):
                    import shutil
                    shutil.copy2(assets_file, full_path)
                    print(f"Файл {filename} скопирован в {full_path}")
                else:
                    raise FileNotFoundError(f"Файл {filename} не найден в assets")
            
            return full_path
            
        except Exception as e:
            print(f"Ошибка получения пути к Excel файлу: {e}")
            return self.original_excel_path
            