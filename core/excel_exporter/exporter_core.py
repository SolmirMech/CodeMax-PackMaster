# exporter_core.py
"""
Ядро системы экспорта в Excel. Только базовые методы.
Обращается к вспомогательным модулям для специфической логики.
"""
import os
from typing import Dict, Any, Optional, List

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from .cell_mappers import SheetMapping, CellMapping, DataType, CellFormat, HorizontalAlignment
from .data_provider import ExportDataProvider
from core.excel_exporter.exporter_data.strategies import get_strategy_for_sheet  # импорт стратегий
from core.excel_exporter.exporter_data.fill_methods import FillMethods  # методы заполнения
from core.excel_exporter.exporter_data.clear_methods import ClearMethods  # методы очистки


class ExcelExportError(Exception):
    """Базовое исключение для ошибок экспорта"""
    pass


class ExcelFileLockedError(ExcelExportError):
    """Файл заблокирован (открыт в Excel)"""
    pass


# noinspection SpellCheckingInspection
class SmartExporter:
    """
    Умный экспортер для работы с Excel.
    Использует DataProvider и CellMappers для декларативного экспорта.
    """
    
    def __init__(self, data_provider: ExportDataProvider, config_manager=None):
        """
        Инициализация экспортера.
        
        Args:
            data_provider: Провайдер данных
            config_manager: Менеджер конфигурации (опционально)
        """
        self.data_provider = data_provider
        self.config_manager = config_manager
        self.wb = None
        self.ws = None
        self.current_mapping: Optional[SheetMapping] = None
        # Инициализация вспомогательных классов
        self.filler = FillMethods(self)      # передаём себя для доступа к wb/ws
        self.cleaner = ClearMethods(self)    # передаём себя для доступа к wb/ws        
        
    # ==================== ОСНОВНЫЕ МЕТОДЫ ЭКСПОРТА ====================

    # noinspection PyUnusedLocal
    # mode передаётся из адаптера и используется там
    def export_to_sheet(self, file_path: str, mapping: SheetMapping,
                       mode: str = "box") -> Dict[str, Any]:
        """Основной метод экспорта в указанный лист."""
        try:
            self._validate_file(file_path)
            self.wb = load_workbook(file_path)
            
            if mapping.sheet_name not in self.wb.sheetnames:
                raise ExcelExportError(f"Лист '{mapping.sheet_name}' не найден")
            
            self.ws = self.wb[mapping.sheet_name]
            self.current_mapping = mapping
            
            # 1. Данные
            all_data = self.data_provider.collect_all_data()
            sheet_data = self._prepare_sheet_data(all_data, mapping.workshop, mapping.sheet_name)
            
            # 2. Статика
            self._fill_static_cells(mapping.static_cells, sheet_data)
            
            # 3. Динамика через стратегию
            strategy = get_strategy_for_sheet(mapping, sheet_data)
            all_fitted = strategy.process(self, mapping, sheet_data)
            
            # 4. Хуки
            self._run_post_processing_hooks(mapping.post_processing_hooks, sheet_data)
            
            self.wb.save(file_path)
            
            return {
                'success': True,
                'all_fitted': all_fitted,
                'file_path': file_path,
                'sheet_name': mapping.sheet_name,
                'workshop': mapping.workshop
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'file_path': file_path,
                'sheet_name': mapping.sheet_name if mapping else 'Unknown'
            }
        finally:
            self._cleanup()
    
    def clear_sheet(self, file_path: str, mapping: SheetMapping) -> bool:
        """Очищает указанный лист."""
        try:
            self._validate_file(file_path, check_lock=True)
            self.wb = load_workbook(file_path)
            
            if mapping.sheet_name not in self.wb.sheetnames:
                return False
            
            self.ws = self.wb[mapping.sheet_name]
            self.current_mapping = mapping
            
            # Очистка через cleaner
            self.cleaner.clear_sheet(mapping)
            
            self.wb.save(file_path)
            return True
        except Exception as e:
            print(f"Ошибка очистки листа: {e}")
            return False
        finally:
            self._cleanup()
            
    # ==================== МЕТОДЫ ЗАПОЛНЕНИЯ ЯЧЕЕК ====================
    
    def _fill_static_cells(self, cell_mappings: List[CellMapping], data: Dict[str, Any]):
        """Заполняет статические ячейки согласно маппингу"""
        for cell_mapping in cell_mappings:
            try:
                # Получаем значение
                value = data.get(cell_mapping.data_key, cell_mapping.default_value)
                
                # Если значение None и ячейка обязательная - пропускаем с предупреждением
                if value is None and cell_mapping.required:
                    print(f"Внимание: отсутствует значение для обязательной ячейки {cell_mapping.cell_reference} "
                          f"(ключ: {cell_mapping.data_key})")
                    continue
                
                # Преобразуем значение согласно типу данных
                processed_value = self.process_value_by_type(value, cell_mapping.data_type)
                
                # Устанавливаем значение в ячейку
                self.set_cell_value(
                    cell_mapping.cell_reference, 
                    processed_value,
                    cell_mapping.format,
                    cell_mapping.is_merged_cell
                )
                
            except Exception as e:
                print(f"Ошибка заполнения ячейки {cell_mapping.cell_reference}: {e}")
                continue
                
    def _prepare_sheet_data(self, all_data: Dict[str, Any], workshop: str, sheet_name: str) -> Dict[str, Any]:
        """
        Подготавливает данные для конкретного листа.
        Можно переопределить в дочерних классах для специфичной логики.
        """
        # Базовый вариант - возвращаем плоскую структуру данных
        sheet_data = {}
        
        # Собираем данные из всех категорий в плоский словарь
        for category, category_data in all_data.items():
            if isinstance(category_data, dict):
                sheet_data.update(category_data)
        
        # Специальная обработка для производителя
        if 'manufacturer' in all_data and isinstance(all_data['manufacturer'], dict):
            manufacturer_data = all_data['manufacturer']
            
            # Переносим display_text под нужным ключом
            if 'display_text' in manufacturer_data:
                sheet_data['manufacturer_display_text'] = manufacturer_data['display_text']
            
            # Также добавляем другие поля производителя
            sheet_data['manufacturer_name'] = manufacturer_data.get('manufacturer_name', '')
            sheet_data['show_manufacturer'] = manufacturer_data.get('show_manufacturer', True)
        
        # Добавляем метаданные
        sheet_data['workshop'] = workshop
        sheet_data['sheet_name'] = sheet_name
        
        # Для поддона 1 цех используем специализированный метод DataProvider
        if sheet_name == "Лист для паллеты" or "паллет" in sheet_name.lower():
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop1_pallet()}
            
        # Для БезВеса 1 цех используем специализированный метод DataProvider
        if sheet_name == "БезВеса":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop1_noweight()}
            
        # Для листа "Много видов 1 цех" используем специализированный метод DataProvider
        if sheet_name == "Лист много видов" and workshop == "1":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop1_multitype()}

        # В методе _prepare_sheet_data, после workshop1_multitype:
        if sheet_name == "Много видов БезВеса" and workshop == "1":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop1_multitype_noweight()}

        if sheet_name == "ПоддонРолики" and workshop == "1":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop1_box_noweight()}

        # Для листа "Много видов 2 цех" используем специализированный метод DataProvider
        if (sheet_name == "Много видов" or "много видов" in sheet_name.lower()) and workshop == "2":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop2_multitype()}
            
        # Для поддона 2 цех (коробка)
        if sheet_name == "Поддон" and workshop == "2":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop2_box()}
            
        # Для списка поддонов 2 цех
        if sheet_name == "Список поддонов" and workshop == "2":
            sheet_data = {**sheet_data, **self.data_provider.get_data_for_workshop2_pallet_list()}
        
        return sheet_data
    
    # ==================== Запуск хуков ====================
    
    def _run_post_processing_hooks(self, hooks: List[str], data: Dict[str, Any]):
        """Выполняет хуки пост-обработки"""
        for hook_name in hooks:
            try:
                method_name = f"_hook_{hook_name}"
                if hasattr(self, method_name):
                    getattr(self, method_name)(data)
                else:
                    print(f"Внимание: хук '{hook_name}' не найден")
            except Exception as e:
                print(f"Ошибка выполнения хука '{hook_name}': {e}")

    def _hook_validate_multitype_noweight_rows(self, data: Dict[str, Any]):
        """
        Хук для проверки количества строк в листе 'Много видов БезВеса'.
        Максимум 18 строк (A11-A28).
        """
        if not self.current_mapping or self.current_mapping.sheet_name != "Много видов БезВеса":
            return

        # Получаем количество заполняемых строк
        boxes_count = data.get('boxes_count', 0)
        max_rows = 18  # строки 11-28

        if boxes_count > max_rows:
            print(
                f"ВНИМАНИЕ: Количество строк ({boxes_count}) превышает максимальное ({max_rows}) для листа 'Много видов БезВеса'")
            data['sheet_overflow'] = True
        elif boxes_count == 0:
            print("ВНИМАНИЕ: Нет данных для заполнения в листе 'Много видов БезВеса'")
                
    def _hook_validate_pallet_list_capacity(self, data: Dict[str, Any]):
        """Хук для проверки заполнения листа 'Список поддонов'"""
        if not self.current_mapping or self.current_mapping.sheet_name != "Список поддонов":
            return
        
        if not self.current_mapping.dynamic_sections:
            return
        
        section = self.current_mapping.dynamic_sections[0]
        start_row, end_row = section.rows_range
        
        # Считаем заполненные строки
        filled_rows = 0
        for row in range(start_row, end_row):
            is_filled = False
            for col in ['D', 'F', 'H', 'L']:
                if self.ws[f'{col}{row}'].value is not None:
                    is_filled = True
                    break
            
            if is_filled:
                filled_rows += 1
        
        # Проверяем, есть ли свободные строки
        free_rows = (end_row - start_row) - filled_rows
        
        if free_rows <= 0:
            print(f"ОШИБКА: Лист 'Список поддонов' переполнен!")
            # Можно добавить флаг в data для возврата ошибки
            data['sheet_overflow'] = True
        elif free_rows <= 3:
            print(f"Внимание: в листе 'Список поддонов' осталось {free_rows} свободных строк")

    @staticmethod
    def _hook_validate_rolls_count_workshop2(data: Dict[str, Any]):
        """Хук для проверки количества роликов для 2 цеха (макс 60)"""
        rolls_count = data.get('rolls_count', 0)
        max_rolls = 60  # Максимум для 2 цеха (3 колонки по 20)
        
        if rolls_count > max_rolls:
            print(f"Внимание: количество роликов ({rolls_count}) превышает максимальное ({max_rolls}) для 2 цеха")
    
    @staticmethod
    def _hook_validate_rolls_count(data: Dict[str, Any]):
        """Хук для проверки количества роликов"""
        rolls_count = data.get('rolls_count', 0)
        max_rolls = 30  # Максимум для 1 цеха (15 слева + 15 справа)
        
        if rolls_count > max_rolls:
            print(f"Внимание: количество роликов ({rolls_count}) превышает максимальное ({max_rolls})")
            
    @staticmethod
    def _hook_validate_boxes_count(data: Dict[str, Any]):
        """Хук для проверки количества коробок"""
        boxes_count = data.get('boxes_count', 0)
        max_boxes = 30  # Максимум для 1 цеха (15 слева + 15 справа)
        
        if boxes_count > max_boxes:
            print(f"Внимание: количество коробок ({boxes_count}) превышает максимальное ({max_boxes})")
            
    @staticmethod
    def _hook_validate_boxes_count_noweight(data: Dict[str, Any]):
        """Хук для проверки количества коробок в листе БезВеса"""
        boxes_count = data.get('boxes_count', 0)
        max_boxes = 45  # Максимум для БезВеса (15 слева + 15 центр + 15 справа)
        
        if boxes_count > max_boxes:
            print(f"Внимание: количество коробок ({boxes_count}) превышает максимальное ({max_boxes}) для листа БезВеса")
            
    def _hook_fill_box_numbers(self, data: Dict[str, Any]):
        """Заполняет номера коробок в листе БезВеса"""
        if self.current_mapping.sheet_name != "БезВеса":
            return
        
        boxes_count = data.get('boxes_count', 0)
        
        # Считаем сколько номеров уже заполнено
        last_number = 0
        
        # Проверяем левые номера (B14-B28)
        for i in range(14, 29):
            value = self.ws[f"B{i}"].value
            if value is not None:
                last_number = max(last_number, int(value))
        
        # Проверяем центральные номера (E14-E28)
        for i in range(14, 29):
            value = self.ws[f"E{i}"].value
            if value is not None:
                last_number = max(last_number, int(value))
        
        # Проверяем правые номера (H14-H28)
        for i in range(14, 29):
            value = self.ws[f"H{i}"].value
            if value is not None:
                last_number = max(last_number, int(value))
        
        # Теперь заполняем пропущенные номера
        current_number = last_number + 1
        boxes_filled = 0
        
        # Левые номера (B14-B28)
        for i in range(14, 29):
            if boxes_filled >= boxes_count:
                break
            if self.ws[f"B{i}"].value is None:
                self.set_cell_value(f"B{i}", current_number, CellFormat(horizontal_alignment=HorizontalAlignment.CENTER))
                current_number += 1
                boxes_filled += 1
        
        # Центральные номера (E14-E28)
        for i in range(14, 29):
            if boxes_filled >= boxes_count:
                break
            if self.ws[f"E{i}"].value is None:
                self.set_cell_value(f"E{i}", current_number, CellFormat(horizontal_alignment=HorizontalAlignment.CENTER))
                current_number += 1
                boxes_filled += 1
        
        # Правые номера (H14-H28)
        for i in range(14, 29):
            if boxes_filled >= boxes_count:
                break
            if self.ws[f"H{i}"].value is None:
                self.set_cell_value(f"H{i}", current_number, CellFormat(horizontal_alignment=HorizontalAlignment.CENTER))
                current_number += 1
                boxes_filled += 1

    # ==================== Вспомогательные методы ====================
    @staticmethod
    def process_value_by_type(value: Any, data_type: DataType) -> Any:
        """Обрабатывает значение согласно типу данных"""
        if value is None:
            return None
        
        try:
            if data_type == DataType.INTEGER:
                if isinstance(value, (int, float)):
                    return int(value)
                elif isinstance(value, str):
                    try:
                        return int(float(value.replace(',', '.')))
                    except:
                        return value
            elif data_type == DataType.NUMBER:
                if isinstance(value, (int, float)):
                    return float(value)
                elif isinstance(value, str):
                    try:
                        return float(value.replace(',', '.'))
                    except:
                        return value
            elif data_type == DataType.DATE:
                # Дата уже должна быть в правильном формате
                return str(value)
            elif data_type == DataType.MULTILINE_TEXT:
                # Многострочный текст
                return str(value)
            elif data_type == DataType.TEXT:
                # Простой текст
                return str(value)
            elif data_type == DataType.FORMULA:
                # Формулы пока не поддерживаем
                return str(value)
            
            # По умолчанию возвращаем как есть
            return value
            
        except Exception as e:
            print(f"Ошибка обработки значения '{value}' как {data_type}: {e}")
            return value
    
    def set_cell_value(self, cell_ref: str, value: Any,
                       format_spec: Optional[CellFormat] = None,
                       is_merged_cell: bool = False):
        """
        Устанавливает значение в ячейку с форматированием.
        Обрабатывает объединенные ячейки.
        """
        try:
            if is_merged_cell:
                # Для объединенных ячеек находим первую ячейку диапазона
                target_cell = self._get_merged_cell_target(cell_ref)
            else:
                target_cell = self.ws[cell_ref]
            
            # Устанавливаем значение
            target_cell.value = value
            
            # Применяем форматирование
            if format_spec:
                alignment = Alignment(
                    horizontal=format_spec.horizontal_alignment.value,
                    vertical=format_spec.vertical_alignment.value,
                    wrap_text=format_spec.wrap_text
                )
                target_cell.alignment = alignment
                
                if format_spec.number_format:
                    target_cell.number_format = format_spec.number_format
                
                if format_spec.bold or format_spec.font_size:
                    font = Font(
                        bold=format_spec.bold,
                        size=format_spec.font_size
                    ) if format_spec.font_size else Font(bold=format_spec.bold)
                    target_cell.font = font
            else:
                # Сбрасываем форматирование
                target_cell.alignment = Alignment(
                    horizontal='general',
                    vertical='center',
                    wrap_text=False
                )
                
        except Exception as e:
            print(f"Ошибка установки значения в ячейку {cell_ref}: {e}")
            raise
    
    def _get_merged_cell_target(self, cell_ref: str):
        """Находит целевую ячейку для объединенного диапазона"""
        cell = self.ws[cell_ref]
        
        # Проверяем, находится ли ячейка в объединенном диапазоне
        for merged_range in self.ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Возвращаем первую ячейку объединенного диапазона
                return self.ws[merged_range.min_row][merged_range.min_col - 1]
        
        # Если не объединенная - возвращаем как есть
        return cell
    
    def _validate_file(self, file_path: str, check_lock: bool = True):
        """Проверяет файл перед работой с ним"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл не найден: {file_path}")
        
        if check_lock and self._is_file_locked(file_path):
            raise ExcelFileLockedError(f"Файл {file_path} открыт в Excel. Закройте его и попробуйте снова.")
    
    @staticmethod
    def _is_file_locked(filepath: str) -> bool:
        """Проверяет, открыт ли файл в другом процессе"""
        try:
            with open(filepath, 'a', encoding='utf-8'):
                pass
            return False
        except (IOError, PermissionError):
            return True
        except Exception:
            return False
    
    def _cleanup(self):
        """Очистка ресурсов"""
        if self.wb:
            try:
                self.wb.close()
            except:
                pass
        self.wb = None
        self.ws = None
        self.current_mapping = None


