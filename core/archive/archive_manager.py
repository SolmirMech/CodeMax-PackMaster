# core/archive/archive_manager.py
import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from core.excel_exporter.cell_mappers import CellMappingRegistry


# noinspection SpellCheckingInspection
class ArchiveManager:
    """Менеджер архива: поиск, восстановление, управление для всех цехов и режимов"""

    def __init__(self, config_manager=None, coordinator=None, exporter=None):
        self.config = config_manager
        self.coordinator = coordinator
        self._exporter = exporter
        self.workshop = "1"  # по умолчанию
        self.has_weight = True  # по умолчанию

        if coordinator and hasattr(coordinator, 'subscribe'):
            coordinator.subscribe(self.on_settings_changed)

        self.on_settings_changed()  # сразу получаем актуальные значения
        self.clean_old_archive_records(3)

    def archive_ecosystem_sheet(self, excel_path):
        """
        Архивирует лист Экосистемы.
        Вызывается напрямую из PackingListWindow.
        """
        try:
            if not excel_path or not os.path.exists(excel_path):
                return {"success": False, "error": f"Файл не найден: {excel_path}"}

            # Получаем маппинг для Экосистемы
            mapping = CellMappingRegistry.get_mapping("1", "ecosystem")

            # Извлекаем данные из переданного файла
            result = self._extract_using_mapping(excel_path, mapping, has_weight=False)

            if result.get("success"):
                archive_data = result.get("archive_data")

                # Добавляем дату в basic_fields в нужном формате
                from datetime import datetime
                archive_data["basic_fields"]["date"] = datetime.now().strftime("%d.%m.%Y")

                # Сохраняем в архив
                archive = self.config.get_pallet_archive()
                if "pallets" not in archive:
                    archive["pallets"] = []
                archive["pallets"].append(archive_data)
                self.config.save_pallet_archive(archive)

                return {"success": True, "message": "Упаковочный лист добавлен в архив"}

            return result

        except Exception as e:
            return {"success": False, "error": str(e)}

    def clean_old_archive_records(self, max_age_years=3):
        """Удаляет записи старше указанного количества лет"""
        try:
            archive = self.config.get_pallet_archive()
            if not archive or "pallets" not in archive:
                return False

            now = datetime.now()
            original_count = len(archive["pallets"])
            archive["pallets"] = [
                pallet for pallet in archive["pallets"]
                if self._is_record_younger_than(pallet, max_age_years, now)
            ]
            new_count = len(archive["pallets"])

            if new_count < original_count:
                self.config.save_pallet_archive(archive)
                print(f"[INFO] Очистка архива: удалено {original_count - new_count} записей старше {max_age_years} лет")

            return True

        except Exception as e:
            print(f"[ERROR] Ошибка очистки архива: {e}")
            return False

    @staticmethod
    def _is_record_younger_than(record, max_age_years, now):
        """Проверяет, что запись не старше max_age_years"""
        extraction_date = record.get("extraction_date")
        if not extraction_date:
            return True

        try:
            record_date = datetime.strptime(extraction_date, "%Y-%m-%d %H:%M:%S")
            age = now - record_date
            return age.days < max_age_years * 365
        except:
            return True

    # noinspection PyUnusedLocal
    def on_settings_changed(self, context=None):
        """Один метод для обновления всех настроек из координатора"""
        if not self.coordinator:
            self.has_weight = False
            self.workshop = "1"
            return

        if hasattr(self.coordinator, 'get_workshop'):
            self.workshop = self.coordinator.get_workshop()

        if hasattr(self.coordinator, 'get_weight_status'):
            self.has_weight = self.coordinator.get_weight_status()

    def get_excel_path(self, workshop=None, has_weight=None):
        """Получает путь к Excel файлу через координатор"""
        # Определяем параметры
        if workshop is None:
            workshop = self.workshop

        if has_weight is None:
            has_weight = self.has_weight

        # Используем координатор (он всегда есть)
        return self.coordinator.get_excel_file_path(workshop, has_weight)

    @staticmethod
    def _get_sheet_info(workshop, enable_pallet, multitype_mode, has_weight=True):
        """Возвращает (имя_листа, тип_листа) на основе параметров"""
        if workshop == "1":
            if multitype_mode:
                # Добавляем проверку веса для мультитайпа
                if not has_weight:
                    return "Много видов БезВеса", "multitype_noweight"
                else:
                    return "Лист много видов", "multitype"
            elif enable_pallet:
                if has_weight:
                    return "Лист для паллеты", "pallet"
                else:
                    return "БезВеса", "noweight"
            else:
                if not has_weight:
                    return "ПоддонРолики", "box_noweight"
                else:
                    return "Лист для коробки", "box"
        else:  # workshop == "2"
            if multitype_mode:
                return "Много видов", "multitype"
            elif enable_pallet:
                return "Список поддонов", "pallet_list"
            else:
                return "Поддон", "box"

    def extract_data_for_archive(self, workshop=None, enable_pallet=False, multitype_mode=False):
        """
        Основной метод архивации с использованием маппингов
        """
        try:
            # Обновляем настройки через существующий метод
            self.on_settings_changed()

            # Определяем цех (если не передан, используем из обновлённых настроек)
            if workshop is None:
                workshop = self.workshop

            # Используем актуальный статус веса из self.has_weight
            has_weight = self.has_weight

            # Определяем путь к файлу
            excel_path = self.get_excel_path(workshop, has_weight)
            if not os.path.exists(excel_path):
                return {"success": False, "error": f"Файл не найден: {excel_path}"}

            # Определяем лист и тип - передаём has_weight
            sheet_name, archive_type = self._get_sheet_info(workshop, enable_pallet, multitype_mode, has_weight)

            # Получаем маппинг
            mapping = CellMappingRegistry.get_mapping(workshop, archive_type)

            # Извлекаем данные через маппинг
            return self._extract_using_mapping(excel_path, mapping, has_weight)

        except Exception as e:
            return {"success": False, "error": str(e)}

    def _extract_using_mapping(self, excel_path, mapping, has_weight=True):
        """Извлекает данные из Excel используя маппинг"""
        workbook = load_workbook(excel_path)
        sheet = workbook[mapping.sheet_name]

        # Статические поля - сохраняем с логическими ключами!
        basic_fields = {}
        for cell_mapping in mapping.static_cells:
            cell_addr = cell_mapping.cell_reference
            # Сохраняем по data_key, а не по адресу ячейки
            basic_fields[cell_mapping.data_key] = sheet[cell_addr].value

        # Динамические секции
        dynamic_data = {}
        for section in mapping.dynamic_sections:
            dynamic_data[section.name] = self._extract_dynamic_section(sheet, section)

        # СПЕЦИАЛЬНАЯ ОБРАБОТКА ДЛЯ ЛИСТА БЕЗВЕСА
        if mapping.sheet_name == "БезВеса":
            # Добавляем номера коробок из колонок B, E, H
            box_numbers = {
                "left": [],
                "center": [],
                "right": []
            }

            # Собираем номера из колонки B
            for row in range(14, 29):
                val = sheet[f"B{row}"].value
                if val is not None:
                    box_numbers["left"].append({"row": row, "value": val})

            # Собираем номера из колонки E
            for row in range(14, 29):
                val = sheet[f"E{row}"].value
                if val is not None:
                    box_numbers["center"].append({"row": row, "value": val})

            # Собираем номера из колонки H
            for row in range(14, 29):
                val = sheet[f"H{row}"].value
                if val is not None:
                    box_numbers["right"].append({"row": row, "value": val})

            # Сохраняем в архив
            dynamic_data["box_numbers"] = box_numbers

        workbook.close()

        return {
            "success": True,
            "archive_data": {
                "workshop": mapping.workshop,
                "archive_type": self._get_type_from_mapping(mapping),
                "sheet_name": mapping.sheet_name,
                "basic_fields": basic_fields,  # Теперь здесь ключи из маппинга!
                "dynamic_data": dynamic_data,
                "has_weight": has_weight,
                "extraction_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        }

    @staticmethod
    def _extract_dynamic_section(sheet, section):
        data = []

        for row_offset in range(section.rows_range[1] - section.rows_range[0] + 1):
            current_row = section.rows_range[0] + row_offset
            row_data = {}
            has_data = False

            # columns_config - это список словарей!
            for col_config in section.columns_config:
                col_letter = col_config["column"]
                data_key = col_config["data_key"]

                cell = f"{col_letter}{current_row}"
                value = sheet[cell].value

                if value is not None:
                    row_data[data_key] = value
                    row_data['_row'] = current_row
                    row_data['_cell'] = cell
                    has_data = True

            if has_data:
                data.append(row_data)

        return data

    # noinspection PyMethodMayBeStatic
    def _get_type_from_mapping(self, mapping):
        """Извлекает тип листа из маппинга"""
        if mapping.sheet_name == "БезВеса":
            return "noweight"
        elif mapping.sheet_name == "Лист для паллеты":
            return "pallet"
        elif mapping.sheet_name == "Лист для коробки":
            return "box"
        elif mapping.sheet_name == "Лист много видов":
            return "multitype"
        elif mapping.sheet_name == "Много видов БезВеса":
            return "multitype_noweight"
        elif mapping.sheet_name == "ПоддонРолики":
            return "box_noweight"
        elif mapping.sheet_name == "Поддон":
            return "box"
        elif mapping.sheet_name == "Список поддонов":
            return "pallet_list"
        elif mapping.sheet_name == "Много видов":
            return "multitype"
        elif mapping.sheet_name == "Экосистема":
            return "ecosystem"
        else:
            return "unknown"

    def restore_to_excel(self, archive_data, excel_path=None):
        """Восстанавливает данные из архива в Excel используя маппинг"""
        try:
            workshop = archive_data.get("workshop", "2")
            archive_type = archive_data.get("archive_type")
            sheet_name = archive_data.get("sheet_name")

            if not archive_type or not sheet_name:
                return {"success": False, "error": "Не указан тип архива или лист"}

            # Получаем маппинг
            mapping = CellMappingRegistry.get_mapping(workshop, archive_type)

            # Определяем путь к файлу, если не передан
            if excel_path is None:
                excel_path = self.get_excel_path(workshop)

            # Очистка перед восстановлением через экспортер
            if self._exporter:
                enable_pallet = (archive_type in ["pallet", "noweight", "pallet_list"])
                multitype_mode = (archive_type == "multitype")
                self._exporter.clear_all_rolls(enable_pallet, multitype_mode)

            # Восстанавливаем через маппинг
            return self._restore_using_mapping(archive_data, excel_path, mapping, workshop)

        except Exception as e:
            return {"success": False, "error": f"Ошибка восстановления: {str(e)}"}

    def _restore_using_mapping(self, archive_data, excel_path, mapping, workshop):
        """Восстанавливает данные используя маппинг"""
        try:
            if mapping.sheet_name == "Экосистема":
                return self._restore_ecosystem(archive_data)

            workbook = load_workbook(excel_path)

            if mapping.sheet_name not in workbook.sheetnames:
                workbook.close()
                return {"success": False, "error": f"Лист '{mapping.sheet_name}' не найден"}

            sheet = workbook[mapping.sheet_name]

            # Словарь значений из архива для быстрого доступа по data_key
            basic_values = archive_data.get("basic_fields", {})

            # Восстанавливаем статические поля - идём ПО МАППИНГУ!
            for cell_mapping in mapping.static_cells:
                # Берём значение по data_key из архива
                value = basic_values.get(cell_mapping.data_key)

                if value is not None:
                    cell_addr = cell_mapping.cell_reference
                    sheet[cell_addr] = value

                    # Применяем форматирование из маппинга
                    if cell_mapping.format.wrap_text:
                        sheet[cell_addr].alignment = Alignment(
                            wrap_text=True
                        )

            # Восстанавливаем динамические секции
            dynamic_data = archive_data.get("dynamic_data", {})

            # Специальное восстановление для листа БезВеса
            if mapping.sheet_name == "БезВеса" and "box_numbers" in dynamic_data:
                box_numbers = dynamic_data["box_numbers"]

                # Восстанавливаем левые номера (B)
                for item in box_numbers.get("left", []):
                    row = item.get("row")
                    if row:
                        sheet[f"B{row}"] = item.get("value")

                # Восстанавливаем центральные номера (E)
                for item in box_numbers.get("center", []):
                    row = item.get("row")
                    if row:
                        sheet[f"E{row}"] = item.get("value")

                # Восстанавливаем правые номера (H)
                for item in box_numbers.get("right", []):
                    row = item.get("row")
                    if row:
                        sheet[f"H{row}"] = item.get("value")

            # Восстанавливаем обычные динамические секции
            sections_by_name = {section.name: section for section in mapping.dynamic_sections}
            for section_name, section_data in dynamic_data.items():
                if section_name in sections_by_name and section_name != "box_numbers":
                    section = sections_by_name[section_name]
                    self._restore_dynamic_section(sheet, section, section_data)

            workbook.save(excel_path)
            workbook.close()

            # Формируем результат
            if workshop == "1":
                # Пробуем получить номер заказа по data_key
                order_num = basic_values.get("order_number", "неизвестно")
            else:
                order_num = basic_values.get("order_number", "неизвестно")

            result = {
                "success": True,
                "sheet_name": mapping.sheet_name,
                "order": order_num
            }

            if workshop == "2":
                pallet_num = basic_values.get("pallet_num", "неизвестно")
                result["pallet_number"] = pallet_num

            return result

        except Exception as e:
            return {"success": False, "error": f"Ошибка восстановления: {str(e)}"}

    @staticmethod
    def _restore_dynamic_section(sheet, section, section_data):
        """Восстанавливает данные в динамическую секцию"""
        for row_data in section_data:
            row = row_data.get('_row')
            if not row:
                continue

            # section.columns_config - это список словарей!
            for col_config in section.columns_config:
                col_letter = col_config["column"]
                data_key = col_config["data_key"]

                if data_key in row_data:
                    cell = f"{col_letter}{row}"
                    sheet[cell] = row_data[data_key]

    def _restore_ecosystem(self, archive_data):
        """
        Восстанавливает лист Экосистемы из архива.
        Использует существующий PackingListExcel.
        """
        try:
            from core.packing_list.packing_list_excel import PackingListExcel

            ecosystem_path = self.config.create_ecosystem_list_work_copy()

            basic_fields = archive_data.get("basic_fields", {})
            dynamic_data = archive_data.get("dynamic_data", {})

            # Собираем данные шапки
            header_data = {
                "list_number": basic_fields.get("list_number", ""),
                "supplier": basic_fields.get("supplier", ""),
                "customer": basic_fields.get("customer", ""),
                "consignee": basic_fields.get("consignee", ""),
                "contract": basic_fields.get("order_number", ""),  # в архиве это order_number
                "project": basic_fields.get("project", ""),
                "equipment_name": basic_fields.get("product_text", ""),  # в архиве это product_text
            }

            # Собираем данные таблицы мест
            places_data = []
            for place in dynamic_data.get("places", []):
                places_data.append({
                    "place_number": place.get("place_number", " "),
                    "net_weight": place.get("net_weight", "0"),
                    "gross_weight": place.get("gross_weight", "0"),
                    "length": place.get("length", "0"),
                    "width": place.get("width", "0"),
                    "height": place.get("height", "0"),
                    "storage_type": place.get("storage_type", " "),
                })

            # Дополняем до 5 строк пустыми значениями
            while len(places_data) < 5:
                places_data.append({
                    "place_number": " ",
                    "net_weight": "0",
                    "gross_weight": "0",
                    "length": "0",
                    "width": "0",
                    "height": "0",
                    "storage_type": " ",
                })

            # Собираем данные таблицы товаров
            items_data = []
            for item in dynamic_data.get("items", []):
                items_data.append({
                    "item_number": item.get("item_number", " "),
                    "order_request": item.get("order_request", " "),
                    "article_vn": item.get("article_vn", " "),
                    "name": item.get("name", " "),
                    "unit": item.get("unit", " "),
                    "quantity": item.get("quantity", "0"),
                    "article_vn_product": item.get("article_vn_product", " "),
                    "product": item.get("product", " "),
                })

            # Дополняем до 5 строк пустыми значениями
            while len(items_data) < 5:
                items_data.append({
                    "item_number": " ",
                    "order_request": " ",
                    "article_vn": " ",
                    "name": " ",
                    "unit": " ",
                    "quantity": "0",
                    "article_vn_product": " ",
                    "product": " ",
                })

            # Заполняем шаблон через PackingListExcel
            PackingListExcel.fill_template(ecosystem_path, header_data, places_data, items_data)

            # Формируем результат
            order_num = basic_fields.get("order_number", "неизвестно")

            return {
                "success": True,
                "sheet_name": "Экосистема",
                "order": order_num
            }

        except Exception as e:
            return {"success": False, "error": f"Ошибка восстановления Экосистемы: {str(e)}"}