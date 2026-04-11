# core/packing_list/packing_list_excel.py

from copy import copy

from openpyxl import load_workbook

from core.packing_list import packing_list_mapping as mapping


class PackingListExcel:
    """Работа с Excel для упаковочного листа Экосистема"""

    @staticmethod
    def fill_template(work_path, header_data, places_data, items_data):
        """
        Открывает копию шаблона, заполняет данными, сохраняет.
        """
        wb = None
        try:
            wb = load_workbook(work_path)
            sheet = wb["Экосистема"]

            # Заполняем шапку (фиксированные ячейки)
            for field, cell_info in mapping.HEADER_MAPPING.items():
                cell = sheet[cell_info["cell"]]
                value = header_data.get(field, "")
                if value is not None:
                    cell.value = value

                style = cell_info.get("style", {})
                if style.get("font"):
                    cell.font = copy(style["font"])
                if style.get("border"):
                    cell.border = copy(style["border"])
                if style.get("alignment"):
                    cell.alignment = copy(style["alignment"])

            # Заполняем таблицу мест
            for i, place in enumerate(places_data):
                if i >= 5:
                    break
                row = mapping.PLACES_START_ROW + i

                for field, col_info in mapping.PLACES_COLUMNS.items():
                    cell = sheet.cell(row=row, column=col_info["col"])
                    value = place.get(field, "")

                    # Числовые поля
                    if field in ["net_weight", "gross_weight", "length", "width", "height"]:
                        if value == "0" or value == "" or value == " ":
                            cell.value = None  # пустая ячейка
                        else:
                            try:
                                cell.value = float(str(value).replace(',', '.'))
                            except:
                                cell.value = value
                    else:
                        # Текстовые поля
                        if value == " ":
                            cell.value = ""
                        else:
                            cell.value = value

                    cell.font = copy(mapping.TABLE_CELL_STYLE["font"])
                    cell.border = copy(mapping.TABLE_CELL_STYLE["border"])
                    cell.alignment = copy(mapping.TABLE_CELL_STYLE["alignment"])

            # Заполняем таблицу товаров
            for i, item in enumerate(items_data):
                if i >= 5:
                    break
                row = mapping.ITEMS_START_ROW + i

                for field, col_info in mapping.ITEMS_COLUMNS.items():
                    cell = sheet.cell(row=row, column=col_info["col"])
                    value = item.get(field, "")

                    # Числовое поле quantity
                    if field == "quantity":
                        if value == "0" or value == "" or value == " ":
                            cell.value = None
                        else:
                            try:
                                cell.value = int(float(str(value).replace(',', '.')))
                            except:
                                cell.value = value
                    else:
                        # Текстовые поля
                        if value == " ":
                            cell.value = ""
                        else:
                            cell.value = value

                    cell.font = copy(mapping.TABLE_CELL_STYLE["font"])
                    cell.border = copy(mapping.TABLE_CELL_STYLE["border"])
                    cell.alignment = copy(mapping.TABLE_CELL_STYLE["alignment"])

            wb.save(work_path)

        except Exception as e:
            raise Exception(f"Ошибка заполнения шаблона: {str(e)}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass
            import gc
            gc.collect()
