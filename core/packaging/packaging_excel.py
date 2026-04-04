# core/packaging/packaging_excel.py
import os
import re
import shutil
import time
from copy import copy

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gc

# noinspection PyUnusedImports
from core.packaging.packaging_mapping import PACKAGING_MAPPINGS, WORKSHOP_1_MAPPING


class PackagingExcel:
    """ВСЯ работа с Excel для журнала упаковки"""

    @staticmethod
    def _mapping_to_import_format(mapping):
        """
        Преобразует PACKAGING_MAPPING в формат для импорта

        Args:
            mapping: словарь маппинга (WORKSHOP_1_MAPPING или WORKSHOP_2_MAPPING)

        Returns:
            dict: {
                "columns": {"field_name": column_number, ...},
                "start_row": int,
                "date_format": str
            }
        """
        return {
            "columns": {
                field: col_map.column
                for field, col_map in mapping["columns"].items()
            },
            "start_row": mapping["start_row"],
            "date_format": mapping.get("date_format", "%d.%m.%Y")
        }

    @staticmethod
    def import_from_excel(file_path, db_callback=None, progress_callback=None,
                          only_first_sheet=True, mapping=None):
        """
        Импорт данных из Excel с динамическим маппингом

        Args:
            file_path: путь к файлу Excel
            db_callback: функция для сохранения записи (entry, sheet_name)
            progress_callback: функция для обновления прогресса (message, count)
            only_first_sheet: импортировать только первый лист
            mapping: маппинг структуры (WORKSHOP_1_MAPPING или WORKSHOP_2_MAPPING)

        Returns:
            tuple: (imported_total, errors)
        """
        errors = []
        imported_total = 0

        # Если маппинг не передан, используем дефолтный (цех 1)
        if mapping is None:
            from core.packaging.packaging_mapping import WORKSHOP_1_MAPPING
            mapping = WORKSHOP_1_MAPPING

        # Преобразуем в формат для импорта
        import_format = PackagingExcel._mapping_to_import_format(mapping)

        try:
            # Используем read_only + keep_links=False для скорости
            wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False, keep_vba=False)

            if only_first_sheet:
                sheets_to_process = [wb.sheetnames[0]] if wb.sheetnames else []
            else:
                sheets_to_process = list(reversed(wb.sheetnames))

            for sheet_idx, sheet_name in enumerate(sheets_to_process):
                sheet = wb[sheet_name]

                if progress_callback:
                    progress_callback(f"Обработка листа {sheet_idx + 1}/{len(sheets_to_process)}: {sheet_name}", None)

                sheet_imported = 0

                # Читаем строки с учётом start_row из маппинга
                start_row = import_format["start_row"]
                for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=False), start_row):
                    # Проверяем первую ячейку (дата) - если пустая, пропускаем
                    first_cell = row[0].value if len(row) > 0 else None
                    if first_cell is None:
                        continue

                    try:
                        entry = PackagingExcel._row_to_entry_from_readonly(row, row_idx, mapping, import_format)
                        if db_callback:
                            db_callback(entry, sheet_name, row_idx)
                        sheet_imported += 1

                        if sheet_imported % 50 == 0:
                            if progress_callback:
                                progress_callback(f"Лист '{sheet_name}'", sheet_imported)
                            gc.collect()

                    except Exception as e:
                        errors.append(f"Лист '{sheet_name}', строка {row_idx}: {str(e)}")

                if sheet_imported > 0:
                    imported_total += sheet_imported
                    if progress_callback:
                        progress_callback(f"✓ Лист '{sheet_name}' завершён", sheet_imported)

                gc.collect()

            wb.close()

            if progress_callback:
                progress_callback("complete", (len(sheets_to_process), imported_total))

        except Exception as e:
            errors.append(f"Ошибка открытия файла: {str(e)}")
            if progress_callback:
                progress_callback("error", str(e))

        return imported_total, errors

    @staticmethod
    def _row_to_entry_from_readonly(row, row_index, mapping, import_format):
        """
        Преобразование строки из read_only режима с динамическим маппингом
        """

        def to_int(value):
            try:
                if isinstance(value, str):
                    value = value.replace(',', '.')
                return int(float(value)) if value is not None else None
            except:
                return None

        def to_float(value):
            try:
                if isinstance(value, str):
                    value = value.replace(',', '.')
                return float(value) if value is not None else None
            except:
                return None

        def format_date(date_cell):
            if not date_cell:
                return ""
            if hasattr(date_cell, 'strftime'):
                return date_cell.strftime('%Y-%m-%d')
            date_match = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', str(date_cell))
            if date_match:
                day, month, year = date_match.groups()
                return f"{year}-{month}-{day}"
            return str(date_cell)[:10]

        # Получаем цвет из первой колонки, которая соответствует col_1..col_N
        row_color = None
        try:
            # Ищем первую колонку из маппинга, которая начинается с 'col_'
            for field, col_map in mapping["columns"].items():
                if field.startswith('col_'):
                    col_index = col_map.column - 1
                    if col_index < len(row):
                        cell = row[col_index]
                        if cell.fill and cell.fill.start_color:
                            color = cell.fill.start_color
                            if hasattr(color, 'rgb') and color.rgb:
                                rgb = color.rgb
                                if len(rgb) == 8:
                                    rgb = rgb[2:]
                                if rgb.upper() not in ('FFFFFF', '000000'):
                                    row_color = rgb
                        break  # берём цвет из первой попавшейся колонки коробок
        except:
            pass

        # Строим запись на основе маппинга
        entry = {'source_row': row_index, 'row_color': row_color}

        for field, column_num in import_format["columns"].items():
            col_index = column_num - 1

            if col_index >= len(row):
                entry[field] = None
                continue

            cell_value = row[col_index].value

            col_mapping = mapping["columns"].get(field)

            if field == "date":
                entry[field] = format_date(cell_value)
            elif col_mapping and col_mapping.data_type == "number":
                entry[field] = to_int(cell_value)
            elif col_mapping and col_mapping.data_type == "float":
                entry[field] = to_float(cell_value)
            else:
                entry[field] = str(cell_value) if cell_value is not None else ""

        return entry

    @staticmethod
    def export_to_excel(file_path, entries, mapping=None, max_retries=3):
        """
        Экспорт записей в Excel с поддержкой обновления существующих строк.
        Возвращает словарь {entry_id: (row_number, sheet_name)} для обновлённых/вставленных записей.
        """
        if not entries:
            return {}

        # Если маппинг не передан, используем дефолтный (цех 1)
        if mapping is None:
            from core.packaging.packaging_mapping import WORKSHOP_1_MAPPING
            mapping = WORKSHOP_1_MAPPING

        lock_file = file_path + ".lock"
        wb = None
        result_coords = {}

        # Блокировка файла
        for attempt in range(max_retries):
            if not os.path.exists(lock_file):
                break
            lock_age = time.time() - os.path.getmtime(lock_file)
            if lock_age > 60:
                try:
                    os.remove(lock_file)
                    break
                except:
                    pass
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                raise Exception("Файл заблокирован другим процессом")

        try:
            with open(lock_file, 'x') as f:
                f.write(str(os.getpid()))
        except FileExistsError:
            raise Exception("Файл заблокирован другим процессом")

        try:
            wb = openpyxl.load_workbook(file_path)
            active_sheet_name = wb.active.title

            # Группируем записи по листам
            entries_by_sheet = {}
            for entry in entries:
                # Для новых записей (без source_sheet) используем активный лист
                sheet_name = entry.get("source_sheet")
                if not sheet_name:
                    sheet_name = active_sheet_name
                    # Сохраняем sheet_name для будущих обновлений
                    result_coords[entry["id"]] = (None, sheet_name)

                if sheet_name not in entries_by_sheet:
                    entries_by_sheet[sheet_name] = []
                entries_by_sheet[sheet_name].append(entry)

            # Обрабатываем каждый лист
            for sheet_name, sheet_entries in entries_by_sheet.items():
                # Проверяем существование листа
                if sheet_name not in wb.sheetnames:
                    # Если листа нет, используем активный (для новых записей)
                    # или игнорируем (старые записи с несуществующим листом)
                    if sheet_name == active_sheet_name:
                        sheet = wb.active
                    else:
                        # Не создаём новый лист, используем активный
                        sheet = wb.active
                        sheet_name = active_sheet_name
                else:
                    sheet = wb[sheet_name]

                # Определяем максимальный номер строки с данными
                max_row = mapping["start_row"] - 1
                for row in range(mapping["start_row"], sheet.max_row + 2):
                    # Проверяем колонку A И колонку B
                    val_a = sheet.cell(row=row, column=1).value
                    val_b = sheet.cell(row=row, column=2).value
                    if val_a is not None or val_b is not None:
                        max_row = row
                    else:
                        break

                # Обрабатываем каждую запись
                for entry in sheet_entries:
                    entry_id = entry["id"]
                    existing_row = entry.get("source_row")
                    existing_sheet = entry.get("source_sheet")

                    # Если запись имеет координаты и они соответствуют текущему листу
                    if existing_row and existing_sheet == sheet_name:
                        row_num = existing_row
                    else:
                        # Новая запись - ищем следующую свободную строку
                        row_num = max_row + 1
                        max_row += 1
                        # Обновляем координаты для новых записей
                        if entry_id in result_coords:
                            result_coords[entry_id] = (row_num, sheet_name)

                    # Заполняем строку
                    row_color = entry.get("row_color")
                    fill = None
                    if row_color:
                        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

                    for field, col_mapping in mapping["columns"].items():
                        col = col_mapping.column
                        value = entry.get(field, "")

                        # Форматируем дату
                        if field == "date" and value and len(str(value)) == 10 and str(value)[4] == '-':
                            y, m, d = str(value).split('-')
                            value = f"{d}.{m}.{y}"

                        # Преобразуем вес в число
                        if field == "weight_kg" and value not in (None, ""):
                            try:
                                value = float(str(value).replace(',', '.'))
                            except:
                                value = value

                        cell = sheet.cell(row=row_num, column=col, value=value)

                        # Применяем стили
                        if col_mapping.style.font:
                            cell.font = copy(col_mapping.style.font)
                        if col_mapping.style.border:
                            cell.border = copy(col_mapping.style.border)
                        if col_mapping.style.alignment:
                            cell.alignment = copy(col_mapping.style.alignment)
                        if col_mapping.style.number_format:
                            cell.number_format = col_mapping.style.number_format

                        # Применяем заливку для колонок коробок
                        if col in (7, 8, 9, 10, 11, 12):
                            if fill:
                                cell.fill = fill
                            else:
                                cell.fill = PatternFill(fill_type=None)

            wb.save(file_path)
            wb.close()
            wb = None

            return result_coords

        except Exception as e:
            raise Exception(f"Ошибка экспорта: {str(e)}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass
            if os.path.exists(lock_file):
                try:
                    os.remove(lock_file)
                except:
                    pass
            import gc
            gc.collect()

    @staticmethod
    def export_entries(entries_by_sheet, template_path, output_path, mapping=None):
        """
        Экспортирует записи в Excel с сохранением структуры листов и цветов строк

        Args:
            entries_by_sheet: список кортежей [(имя_листа, [список_записей]), ...]
            template_path: путь к файлу-шаблону
            output_path: путь для сохранения результата
            mapping: маппинг структуры (WORKSHOP_1_MAPPING или WORKSHOP_2_MAPPING)

        Returns:
            int: количество экспортированных записей
        """
        # Если маппинг не передан, используем дефолтный (цех 1)
        if mapping is None:
            from core.packaging.packaging_mapping import WORKSHOP_1_MAPPING
            mapping = WORKSHOP_1_MAPPING

        wb = None
        try:
            # Копируем шаблон
            shutil.copy2(template_path, output_path)

            # Открываем копию
            wb = load_workbook(output_path)

            start_row = mapping["start_row"]

            total_exported = 0

            # Для каждой группы создаём свой лист
            for sheet_idx, (sheet_name, entries) in enumerate(entries_by_sheet):
                if not entries:
                    continue

                # Копируем структуру из листа-шаблона (который называется "Шаблон")
                ws = wb.copy_worksheet(wb['Шаблон'])
                ws.title = sheet_name[:31]
                # Закрепить заголовок
                if mapping.get("name") == "Цех 2":
                    ws.freeze_panes = 'A3'
                else:
                    ws.freeze_panes = 'A2'

                # Заполняем данными
                for idx, entry in enumerate(entries):
                    row_num = start_row + idx

                    # Получаем цвет строки
                    row_color = entry.get("row_color")
                    fill = None
                    if row_color:
                        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

                    for field, col_map in mapping["columns"].items():
                        value = entry.get(field, "")

                        # Форматируем дату
                        if field == "date" and value and len(str(value)) == 10 and str(value)[4] == '-':
                            y, m, d = str(value).split('-')
                            value = f"{d}.{m}.{y}"

                        cell = ws.cell(row=row_num, column=col_map.column)

                        # Проверяем, не входит ли ячейка в объединённый диапазон
                        skip_cell = False
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                skip_cell = True
                                break

                        if skip_cell:
                            continue

                        cell.value = value

                        # Применяем стили из маппинга
                        if col_map.style.font:
                            cell.font = copy(col_map.style.font)
                        if col_map.style.border:
                            cell.border = copy(col_map.style.border)
                        if col_map.style.alignment:
                            cell.alignment = copy(col_map.style.alignment)
                        if col_map.style.number_format:
                            cell.number_format = col_map.style.number_format

                        # Применяем заливку для всех колонок, если есть цвет строки
                        # Если нужно ограничить только колонками с коробками (7-12)
                        if fill and col_map.column in (7, 8, 9, 10, 11, 12):
                            cell.fill = fill

                total_exported += len(entries)

            # Сохраняем и закрываем
            wb.save(output_path)
            wb.close()
            wb = None

            # Принудительная сборка мусора для освобождения файла
            import gc
            gc.collect()

            return total_exported

        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            return 0
        finally:
            # Гарантированное закрытие в случае ошибки
            if wb:
                try:
                    wb.close()
                except:
                    pass
