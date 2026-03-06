# core/packaging/packaging_data_manager.py
import re
import sqlite3
import threading

import openpyxl


class PackagingDataManager:
    """Кэширующий слой БД для журнала упаковки (аналог XMLDataManager)"""

    def __init__(self, config_manager, coordinator=None):
        """
        Инициализация через ConfigManager.

        Args:
            config_manager: Экземпляр ConfigManager для получения путей
            coordinator: Координатор для подписки на уведомления
        """
        self.status_callback = None
        self.config = config_manager
        self.coordinator = coordinator

        # Путь к БД в папке data AppData
        self.db_path = config_manager.data_dir / "packaging.db"

        # Атрибуты для управления потоками (КОПИРУЕМ из XMLDataManager)
        self._readers_count = 0  # Счётчик активных читателей
        self._readers_lock = threading.RLock()  # Блокировка для счётчика
        self._write_lock = threading.RLock()  # Блокировка для записи
        self._lock = threading.RLock()  # Основная блокировка

        # Инициализация БД
        self._init_database()

        # Подписываемся на уведомления координатора
        if self.coordinator and hasattr(self.coordinator, 'subscribe'):
            self.coordinator.subscribe(self.on_settings_changed)
        
    def on_settings_changed(self, context=None):
        pass
        
    def set_status_callback(self, callback):
        """Устанавливает callback для отправки статусных сообщений в UI"""
        self.status_callback = callback

    def _init_database(self):
        """Инициализация БД для журнала упаковки"""
        with self._lock:
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                # Создаём таблицу packaging_log
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS packaging_log (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        date TEXT,
                        order_number TEXT,
                        customer TEXT,
                        product_name TEXT,
                        quantity_labels INTEGER DEFAULT 0,
                        packer_name TEXT,
                        large_boxes INTEGER DEFAULT 0,
                        small_boxes INTEGER DEFAULT 0,
                        aquaLife_boxes INTEGER DEFAULT 0,
                        note TEXT
                    )
                """)

                # Индексы для поиска
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_pack_date ON packaging_log(date)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_pack_order ON packaging_log(order_number)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_pack_customer ON packaging_log(customer)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_pack_product ON packaging_log(product_name)")

                conn.commit()

            except sqlite3.Error as e:
                print(f"Ошибка инициализации БД: {e}")
            finally:
                if 'conn' in locals():
                    conn.close()

    def get_recent(self, limit=10):
        """Последние записи - копируем блокировки и подход из XMLDataManager"""
        with self._readers_lock:
            self._readers_count += 1
        try:
            with self._lock:
                conn = sqlite3.connect(self.db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT * FROM packaging_log 
                    ORDER BY id DESC 
                    LIMIT ?
                """, (limit,))
                rows = cursor.fetchall()
                return [dict(row) for row in rows]
        finally:
            with self._readers_lock:
                self._readers_count -= 1
            conn.close()

    def search(self, filters):
        """Поиск по одному или нескольким полям - копируем подход с блокировками"""
        with self._readers_lock:
            self._readers_count += 1
        try:
            with self._lock:
                conn = sqlite3.connect(self.db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                query = "SELECT * FROM packaging_log WHERE 1=1"
                params = []

                for key, value in filters.items():
                    if key == 'date':
                        query += " AND date = ?"
                        params.append(value)
                    elif key in ['order_number', 'customer', 'product_name']:
                        query += f" AND {key} LIKE ?"
                        params.append(f'%{value}%')

                query += " ORDER BY id DESC"

                cursor.execute(query, params)
                rows = cursor.fetchall()
                return [dict(row) for row in rows]
        finally:
            with self._readers_lock:
                self._readers_count -= 1
            conn.close()

    def update_entry(self, entry_id, field, value):
        """Обновление конкретной ячейки - с блокировкой записи"""
        with self._write_lock:
            try:
                with self._lock:
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute(f"UPDATE packaging_log SET {field} = ? WHERE id = ?", (value, entry_id))
                    conn.commit()
                    return cursor.rowcount > 0
            finally:
                conn.close()

    def add_entry(self, data):
        """Добавление новой записи - с блокировкой записи"""
        with self._write_lock:
            try:
                with self._lock:
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute("""
                        INSERT INTO packaging_log 
                        (date, order_number, customer, product_name, quantity_labels, 
                         packer_name, large_boxes, small_boxes, aquaLife_boxes, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        data.get('date', ''),
                        data.get('order_number', ''),
                        data.get('customer', ''),
                        data.get('product_name', ''),
                        data.get('quantity_labels') if data.get('quantity_labels') not in (None, '') else None,
                        data.get('packer_name', ''),
                        data.get('large_boxes') if data.get('large_boxes') not in (None, '') else None,
                        data.get('small_boxes') if data.get('small_boxes') not in (None, '') else None,
                        data.get('aquaLife_boxes') if data.get('aquaLife_boxes') not in (None, '') else None,
                        data.get('note', '')
                    ))
                    conn.commit()
                    return cursor.lastrowid
            finally:
                conn.close()

    def delete_entry(self, entry_id):
        """Удаление записи - с блокировкой записи"""
        with self._write_lock:
            try:
                with self._lock:
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM packaging_log WHERE id = ?", (entry_id,))
                    conn.commit()
                    return cursor.rowcount > 0
            finally:
                conn.close()

    def import_from_excel(self, file_path, sheet_name="янв 26"):
        """
        Импорт данных из Excel файла.

        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа с данными (по умолчанию "янв 26")

        Returns:
            tuple: (количество импортированных записей, список ошибок)
        """
        errors = []
        imported = 0

        with self._write_lock:
            conn = None
            try:
                # Загружаем книгу
                wb = openpyxl.load_workbook(file_path, data_only=True)

                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return 0, [f"Лист '{sheet_name}' не найден в файле"]

                sheet = wb[sheet_name]

                # Ищем заголовки в первой строке
                headers = {}
                for col_idx, cell in enumerate(next(sheet.iter_rows(max_row=1)), 1):
                    if cell.value and isinstance(cell.value, str):
                        header = cell.value.strip().lower()
                        if header in ["дата", "№ заказа", "заказчик", "наименование",
                                      "тираж", "упаковщик", "большие", "маленькие",
                                      "аквалайф", "примечание"]:
                            headers[header] = col_idx

                # Проверяем наличие обязательных колонок
                required = ["дата", "№ заказа"]
                missing = [r for r in required if r not in headers]
                if missing:
                    wb.close()
                    return 0, [f"Не найдены обязательные колонки: {missing}"]

                # Одно соединение на весь импорт
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                # Проходим по строкам начиная со 2-й
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                    try:
                        # Пропускаем полностью пустые строки
                        if all(cell.value is None for cell in row):
                            continue

                        data = {}

                        # Дата
                        if "дата" in headers:
                            date_cell = row[headers["дата"] - 1].value
                            if date_cell:
                                if hasattr(date_cell, 'strftime'):
                                    data['date'] = date_cell.strftime('%d.%m.%Y')
                                else:
                                    date_match = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', str(date_cell))
                                    if date_match:
                                        day, month, year = date_match.groups()
                                        data['date'] = f"{day}.{month}.{year}"
                                    else:
                                        data['date'] = str(date_cell)[:10]

                        # Номер заказа
                        if "№ заказа" in headers:
                            data['order_number'] = str(row[headers["№ заказа"] - 1].value or "")

                        # Заказчик
                        if "заказчик" in headers:
                            data['customer'] = str(row[headers["заказчик"] - 1].value or "")

                        # Наименование
                        if "наименование" in headers:
                            data['product_name'] = str(row[headers["наименование"] - 1].value or "")

                        # Тираж
                        if "тираж" in headers:
                            val = row[headers["тираж"] - 1].value
                            try:
                                data['quantity_labels'] = int(float(val)) if val else None
                            except (ValueError, TypeError):
                                data['quantity_labels'] = None

                        # Упаковщик
                        if "упаковщик" in headers:
                            data['packer_name'] = str(row[headers["упаковщик"] - 1].value or "")

                        # Большие коробки
                        if "большие" in headers:
                            val = row[headers["большие"] - 1].value
                            try:
                                data['large_boxes'] = int(float(val)) if val else None
                            except (ValueError, TypeError):
                                data['large_boxes'] = None

                        # Маленькие коробки
                        if "маленькие" in headers:
                            val = row[headers["маленькие"] - 1].value
                            try:
                                data['small_boxes'] = int(float(val)) if val else None
                            except (ValueError, TypeError):
                                data['small_boxes'] = None

                        # Аквалайф
                        if "аквалайф" in headers:
                            val = row[headers["аквалайф"] - 1].value
                            try:
                                data['aquaLife_boxes'] = int(float(val)) if val else None
                            except (ValueError, TypeError):
                                data['aquaLife_boxes'] = None

                        # Примечание
                        if "примечание" in headers:
                            data['note'] = str(row[headers["примечание"] - 1].value or "")

                        # Проверяем, что есть хоть какие-то данные
                        if data.get('order_number') or data.get('date'):
                            cursor.execute("""
                                INSERT INTO packaging_log 
                                (date, order_number, customer, product_name, quantity_labels, 
                                 packer_name, large_boxes, small_boxes, aquaLife_boxes, note)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                data.get('date', ''),
                                data.get('order_number', ''),
                                data.get('customer', ''),
                                data.get('product_name', ''),
                                data.get('quantity_labels'),
                                data.get('packer_name', ''),
                                data.get('large_boxes'),
                                data.get('small_boxes'),
                                data.get('aquaLife_boxes'),
                                data.get('note', '')
                            ))
                            imported += 1

                        # Коммитим каждые 50 строк
                        if imported % 50 == 0:
                            conn.commit()

                    except Exception as e:
                        errors.append(f"Строка {row_idx}: {str(e)}")

                conn.commit()
                wb.close()

            except Exception as e:
                errors.append(f"Ошибка при импорте: {str(e)}")
            finally:
                if conn:
                    conn.close()

        return imported, errors
